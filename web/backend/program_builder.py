"""
Fills a .xlsm program template with patient data using openpyxl,
then converts to PDF via LibreOffice headless.
No Excel/xlwings required — runs on Linux (Render.com).
"""
import io
import os
import re
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", ".."))

TEMPLATE_MAP = {
    ("Body Masters", "upper"): os.path.join(WORKSPACE_DIR, "Body Masters", "Body Masters Upper Body.xlsm"),
    ("Body Masters", "lower"): os.path.join(WORKSPACE_DIR, "Body Masters", "Body Masters Lower Body.xlsm"),
    ("Body Masters", "full"):  os.path.join(WORKSPACE_DIR, "Body Masters", "Body Masters Full Body.xlsm"),
    ("Body Motions", "upper"): os.path.join(WORKSPACE_DIR, "Body Motions", "Body Motions Upper Body.xlsm"),
    ("Body Motions", "lower"): os.path.join(WORKSPACE_DIR, "Body Motions", "Body Motions Lower Body.xlsm"),
    ("Body Motions", "full"):  os.path.join(WORKSPACE_DIR, "Body Motions", "Body Motions Full Body.xlsm"),
}

BODY_PART_MAP = {
    "shoulder":      "Shoulder / الكتف",
    "rotator_cuff":  "Rotator cuff/ الكُمّ",
    "chest":         "Chest/ الصدر",
    "back":          "Back/ الظهر",
    "triceps":       "Triceps /التراي سيبس",
    "biceps":        "Biceps /الباي سيبس",
    "hand":          "Hand/ اليد",
    "quadriceps":    "Quadriceps / الفخذ الأمامي",
    "hamstring":     "Hamstring / الفخذ الخلفي",
    "hip_abductors": "Hip Abductors / الفخذ  الخارجي",
    "hip_adductors": "Hip Adductors / الفخذ الداخلي",
    "hip_flexors":   "Hip Flexors / عضلات ثني الورك",
    "hip_extensors": "Hip Extensors/عضلات بسط الورك",
    "trunk":         "Trunk / الجذع",
}

ORDERED_PARTS = {
    "upper": ["shoulder", "rotator_cuff", "chest", "back", "triceps", "biceps", "hand"],
    "lower": ["quadriceps", "hamstring", "hip_abductors", "hip_adductors", "hip_flexors", "hip_extensors", "trunk"],
    "full":  ["shoulder", "rotator_cuff", "triceps", "biceps", "hamstring", "quadriceps", "hip_abductors", "hip_adductors"],
}

PERCENTAGE_REMARK_PAIRS = [
    ("D21", "D22"), ("D23", "D24"), ("D25", "D26"), ("D27", "D28"),
    ("P21", "P22"), ("P23", "P24"), ("P25", "P26"), ("P27", "P28"),
    ("AB21", "AB22"), ("AB23", "AB24"), ("AB25", "AB26"), ("AB27", "AB28"),
    ("AH21", "AH22"), ("AH23", "AH24"), ("AH25", "AH26"), ("AH27", "AH28"),
]


def _get_remark(pct_fraction):
    if pct_fraction is None:
        return ""
    pct = abs(pct_fraction) * 100
    if 0.1 <= pct <= 3.9:   return "Perfect Symmetry / \nتناظر مثالي"
    elif 4 <= pct <= 7.9:   return "Normal Symmetry / \nتناظر طبيعي"
    elif 8 <= pct <= 14.9:  return "Weakness / \nضعف"
    elif 15 <= pct <= 19.9: return "Problem / \nمشكلة"
    elif 20 <= pct <= 29.9: return "Major Problem / \nمشكلة كبيرة"
    elif pct >= 30:         return "Risk Of Injury / \nخطر الإصابة"
    return ""


def _body_parts_present(movements, test_type):
    """Given movements list of [movement, region], return set of body part keys."""
    present = set()
    for movement, region in movements:
        m, r = movement.lower().strip(), region.lower().strip()
        if test_type in ("upper", "full"):
            if r == "shoulder":
                if m in ("external rotation", "internal rotation"):
                    present.add("rotator_cuff")
                elif m in ("flexion", "abduction"):
                    present.add("shoulder")
                elif m == "push":
                    present.add("chest")
                elif m == "pull":
                    present.add("back")
            elif r == "elbow":
                present.add("triceps" if m == "extension" else "biceps")
            elif r == "hand":
                present.add("hand")
        if test_type in ("lower", "full"):
            if r == "knee":
                present.add("quadriceps" if m == "extension" else "hamstring")
            elif r == "hip":
                if m == "adduction":   present.add("hip_adductors")
                elif m == "abduction": present.add("hip_abductors")
                elif m == "flexion":   present.add("hip_flexors")
                elif m == "extension": present.add("hip_extensors")
            elif r == "trunk":
                present.add("trunk")
    return present


def generate_program_xlsm(gym: str, test_type: str, patient_name: str,
                           test_date: str, cells_data: dict) -> bytes:
    """
    Fill the .xlsm template for a patient and return the file as bytes.

    cells_data: dict with keys:
        "cells"     - {cell_addr: value} from check_processor
        "movements" - [[movement, region], ...] list
    """
    template_path = TEMPLATE_MAP.get((gym, test_type))
    if not template_path or not os.path.isfile(template_path):
        raise FileNotFoundError(f"Template not found for {gym} / {test_type} at {template_path}")

    wb = load_workbook(template_path, keep_vba=True)

    # Find the sheet (named after the gym)
    ws = None
    for sheet in wb.sheetnames:
        if sheet.lower() == gym.lower():
            ws = wb[sheet]
            break
    if ws is None:
        ws = wb.active  # fallback

    # Patient name → A6
    normalized_name = re.sub(r"\s+", " ", patient_name.strip())
    ws["A6"] = normalized_name

    # Test date → A21 (display as DD/MM/YYYY)
    if test_date:
        try:
            d = datetime.strptime(str(test_date)[:10], "%Y-%m-%d")
            ws["A21"] = d.strftime("%d/%m/%Y")
        except Exception:
            ws["A21"] = test_date

    # Fill all asymmetry cells
    cells = cells_data.get("cells", {})
    for cell_addr, value in cells.items():
        ws[cell_addr] = value

    # Remarks for percentage cells
    for pct_cell, remark_cell in PERCENTAGE_REMARK_PAIRS:
        pct_value = ws[pct_cell].value
        if pct_value is not None and pct_value != "":
            remark = _get_remark(pct_value)
            if remark:
                ws[remark_cell] = remark

    # Body parts list → A11–A18
    movements = cells_data.get("movements", [])
    present_parts = _body_parts_present(movements, test_type)
    ordered = ORDERED_PARTS.get(test_type, [])
    parts_to_write = [BODY_PART_MAP[p] for p in ordered if p in present_parts]
    for i, label in enumerate(parts_to_write[:8]):
        ws[f"A{11 + i}"] = label

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF generation — HTML rendered via weasyprint (no LibreOffice needed)
# ---------------------------------------------------------------------------

# All (label_cell, pct_cell, side_cell) triplets in template order
_MOVEMENT_CELL_GROUPS = []
for _pct_col, _lbl_col in [("D", "C"), ("P", "O"), ("AB", "AA"), ("AH", "AG")]:
    for _row in range(21, 29, 2):
        _movement_cell_groups_entry = (_lbl_col + str(_row), _pct_col + str(_row), _lbl_col + str(_row + 1))
        _MOVEMENT_CELL_GROUPS.append(_movement_cell_groups_entry)

_REMARK_COLOR = {
    "Perfect Symmetry": "#16a34a",
    "Normal Symmetry":  "#22c55e",
    "Weakness":         "#ca8a04",
    "Problem":          "#ea580c",
    "Major Problem":    "#dc2626",
    "Risk Of Injury":   "#991b1b",
}

def _remark_short(pct_fraction):
    """Return the English part of the remark only."""
    full = _get_remark(pct_fraction)
    return full.split(" /")[0].strip() if full else ""

def _remark_color(pct_fraction):
    short = _remark_short(pct_fraction)
    for key, color in _REMARK_COLOR.items():
        if short.startswith(key):
            return color
    return "#6b7280"

def _parse_movement_rows(cells: dict):
    """Extract (label, pct_value, side_label) from the cells dict."""
    rows = []
    for lbl_cell, pct_cell, side_cell in _MOVEMENT_CELL_GROUPS:
        pct = cells.get(pct_cell)
        if pct is None:
            continue
        label = str(cells.get(lbl_cell, "")).split("\n")[0].strip()
        side  = str(cells.get(side_cell, "")).split("\n")[0].strip()
        rows.append((label, float(pct), side))
    return rows


def generate_program_pdf(gym: str, test_type: str, patient_name: str,
                         test_date: str, cells_data: dict):
    """
    Generate a filled program PDF using weasyprint (pure Python, no LibreOffice).
    Returns (bytes, content_type, filename).
    """
    from weasyprint import HTML as WeasyprintHTML

    safe_name = re.sub(r'[^\w\s-]', '', patient_name).strip().replace(' ', '_')
    type_label = {"upper": "Upper Body", "lower": "Lower Body", "full": "Full Body"}.get(test_type, test_type)
    base_filename = f"{safe_name}_-_{type_label.replace(' ', '_')}"

    # Format date nicely
    display_date = test_date
    try:
        display_date = datetime.strptime(str(test_date)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        pass

    # Parse movement rows from cells
    cells = cells_data.get("cells", {})
    movement_rows = _parse_movement_rows(cells)

    # Build table rows HTML
    table_rows_html = ""
    for label, pct, side in movement_rows:
        pct_display = f"{abs(pct) * 100:.1f}%"
        remark = _remark_short(pct)
        color = _remark_color(pct)
        table_rows_html += f"""
        <tr>
          <td class="label">{label}</td>
          <td class="pct">{pct_display}</td>
          <td class="side">{side}</td>
          <td class="remark" style="color:{color};font-weight:600">{remark}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
  @page {{ size: A4; margin: 18mm 14mm; }}
  body {{ font-family: Arial, sans-serif; font-size: 11pt; color: #111; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-start;
             border-bottom: 3px solid #1d4ed8; padding-bottom: 10px; margin-bottom: 18px; }}
  .gym {{ font-size: 20pt; font-weight: 700; color: #1d4ed8; }}
  .meta {{ text-align: right; font-size: 10pt; color: #444; line-height: 1.7; }}
  .meta strong {{ color: #111; font-size: 12pt; }}
  h2 {{ font-size: 13pt; color: #1e40af; margin: 0 0 10px 0; text-transform: uppercase;
        letter-spacing: .05em; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 4px; }}
  th {{ background: #1d4ed8; color: #fff; padding: 7px 10px; text-align: left;
        font-size: 10pt; }}
  td {{ padding: 6px 10px; border-bottom: 1px solid #e5e7eb; font-size: 10pt; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  .pct {{ font-weight: 700; text-align: center; width: 70px; }}
  .side {{ color: #555; width: 160px; }}
  .remark {{ width: 160px; }}
  .label {{ }}
  .footer {{ margin-top: 22px; font-size: 9pt; color: #9ca3af; text-align: center;
             border-top: 1px solid #e5e7eb; padding-top: 8px; }}
</style>
</head>
<body>
  <div class="header">
    <div class="gym">{gym}</div>
    <div class="meta">
      <strong>{patient_name}</strong><br>
      {type_label} Program<br>
      Date: {display_date}
    </div>
  </div>

  <h2>Asymmetry Results</h2>
  <table>
    <thead>
      <tr>
        <th>Movement</th>
        <th style="text-align:center">Asymmetry</th>
        <th>Weaker Side</th>
        <th>Status</th>
      </tr>
    </thead>
    <tbody>
      {table_rows_html}
    </tbody>
  </table>

  <div class="footer">Generated by VALD Automator · {gym} · {display_date}</div>
</body>
</html>"""

    pdf_bytes = WeasyprintHTML(string=html).write_pdf()
    return pdf_bytes, "application/pdf", f"{base_filename}.pdf"
