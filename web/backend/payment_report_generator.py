"""
Generates the payment report by appending a new month's programs
to the cumulative Payment - Month YEAR.xlsx template.

Structure of the payment file:
  - REPORT sheet: summary of totals per branch (col B=Masters, col E=Motions)
  - One sheet per branch named like "Body Motions - RUH - Al Malaz"
    Each branch sheet:
      row 3:  B3 = "NUMBER OF PROGRAMS:" | C3 = =COUNTA(E7:E13475)  (auto-updates)
      row 5-6: bilingual headers
      row 7+:  data rows: col A=client_id, B=client_name, C=trainer_name, D=test_date, E=dispatch_date
    Green separator row (fill FF8CC075) inserted between months.
"""
import io
import os
import re
from datetime import date, datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

BASE_DIR = os.path.dirname(__file__)
PAYMENT_TEMPLATE_PATH = os.path.join(BASE_DIR, "Payment - Month YEAR.xlsx")

# ── Green separator fill ───────────────────────────────────────────────────────
GREEN_FILL = PatternFill(fill_type="solid", fgColor="8CC075")

# ── Maps each payment sheet name → (gym, branch key as stored in Supabase) ───
PAYMENT_SHEET_TO_BRANCH: dict[str, tuple[str, str]] = {
    # Body Motions
    "Body Motions - RUH - Al Malaz":    ("Body Motions", "RUH - Al Malaz"),
    "Body Motions - RUH - Al Sahafa":   ("Body Motions", "RUH - Al Sahafa"),
    "Body Motions - RUH - Al Aarid":    ("Body Motions", "RUH - Al Aarid"),
    "Body Motions - RUH - Al Fayha":    ("Body Motions", "RUH - Al Fayha"),
    "Body Motions - RUH - Al Uraija":   ("Body Motions", "RUH - Al Uraija"),
    "Body Motions - RUH - Badr":        ("Body Motions", "RUH - Badr"),
    "Body Motions - RUH - Al Badia":    ("Body Motions", "RUH - Al Badia"),
    "Body Motions-JED-Al Basateen":     ("Body Motions", "JED - Al Basateen"),
    "Body Motions-JED-Al Faisaliyah":   ("Body Motions", "JED - Al Faisaliyah"),
    "Body Motions - JED - Al Naeem":    ("Body Motions", "JED - Al Naeem"),
    "Body Motions-DMM-Al Faisaliyah":   ("Body Motions", "DMM - Al Faisaliyah"),
    "Body Motions -DMM- Al Jalawiah":   ("Body Motions", "DMM - Al Jalawiah"),
    "Body Motions - DMM - Al Nada":     ("Body Motions", "DMM - Al Nada"),
    "Body Motions - ALQ - Buraidah":    ("Body Motions", "ALQ - Buraidah"),
    "Body Motions - ALQ - Unaizah":     ("Body Motions", "ALQ - Unaizah"),
    "Body Motions - Al Ahsaa":          ("Body Motions", "Al Ahsaa"),
    "Body Motions - AlUla":             ("Body Motions", "AlUla"),
    "Body Motions - Tabuk":             ("Body Motions", "Tabuk"),
    # Body Masters
    "Body Masters - RUH - Al Malaz":    ("Body Masters", "RUH - Al Malaz"),
    "Body Masters - RUH - Al Massif":   ("Body Masters", "RUH - Al Massif"),
    "Body Masters - RUH - Al Aarid":    ("Body Masters", "RUH - Al Aarid"),
    "Body Masters - RUH - Al Sahafa":   ("Body Masters", "RUH - Al Sahafa"),
    "Body Masters - RUH - Al Wadi":     ("Body Masters", "RUH - Al Wadi"),
    "Body Masters - RUH - Eshbilia":    ("Body Masters", "RUH - Eshbilia"),
    "Body Masters - RUH - Muzahmiyah":  ("Body Masters", "RUH - Muzahmiyah"),
    "Body Masters - RUH - Rabwa":       ("Body Masters", "RUH - Rabwa"),
    "Body Masters - RUH - Salam":       ("Body Masters", "RUH - Salam"),
    "Body Masters - RUH - Swaidi":      ("Body Masters", "RUH - Swaidi"),
    "Body Masters - RUH - Takhasousi":  ("Body Masters", "RUH - Takhasousi"),
    "Body Masters - RUH - Al Badia":    ("Body Masters", "RUH - Al Badia"),
    "Body Masters - RUH - Al Fayha":    ("Body Masters", "RUH - Al Fayha"),
    "Body Masters - RUH - Al Khaleej":  ("Body Masters", "RUH - Al Khaleej"),
    "Body Masters - RUH - Al Kharj":    ("Body Masters", "RUH - Al Kharj"),
    "Body Masters - RUH - Al Nahda":    ("Body Masters", "RUH - Al Nahda"),
    "Body Masters - RUH - Badr":        ("Body Masters", "RUH - Badr"),
    "Body Masters - RUH - Ezdehar":     ("Body Masters", "RUH - Ezdehar"),
    "Body Masters - RUH - Murooj":      ("Body Masters", "RUH - Murooj"),
    "Body Masters - RUH - Shubra":      ("Body Masters", "RUH - Shubra"),
    "Body Masters - DMM - Al Athir":    ("Body Masters", "DMM - Al Athir"),
    "Body Masters-DMM-Al Jameyeen":     ("Body Masters", "DMM - Al Jameyeen"),
    "Body Masters - DMM - Hufof":       ("Body Masters", "DMM - Hufof"),
    "Body Masters - DMM - Khobar":      ("Body Masters", "DMM - Khobar"),
    "Body Masters - JED - Hamadania":   ("Body Masters", "JED - Hamadania"),
    "Body Masters - JED - Al Rawdah":   ("Body Masters", "JED - Al Rawdah"),
    "Body Masters - JED - Makkah":      ("Body Masters", "JED - Makkah"),
    "Masters -JED- Obhor - Al Amwaj":   ("Body Masters", "JED - Obhor - Al Amwaj"),
    "Masters -JED- Obhor - Al Sheraa":  ("Body Masters", "JED - Obhor - Al Sheraa"),
    "Body Masters - ALQ - Al Rass":     ("Body Masters", "ALQ - Al Rass"),
    "Body Masters - ALQ - Buraidah":    ("Body Masters", "ALQ - Buraidah"),
    "Body Masters - ALQ - Unaizah":     ("Body Masters", "ALQ - Unaizah"),
    "Body Masters - MED - Shouran":     ("Body Masters", "MED - Shouran"),
    "Body Masters - MED - Taiba":       ("Body Masters", "MED - Taiba"),
    "Body Masters - Uhud":              ("Body Masters", "Uhud"),
    "Body Masters - AlUla":             ("Body Masters", "AlUla"),
    "Body Masters - Al Mubaraz":        ("Body Masters", "Al Mubaraz"),
    "Body Masters - Hafr El Batin":     ("Body Masters", "Hafr El Batin"),
    "Body Masters - Tabuk":             ("Body Masters", "Tabuk"),
    "Body Masters - Najran":            ("Body Masters", "Najran"),
    "Body Masters - Khamis Mushait":    ("Body Masters", "Khamis Mushait"),
    "Body Masters - Hail":              ("Body Masters", "Hail"),
}

# ── REPORT sheet: maps REPORT branch label → list of payment sheet names ─────
# Used to count cumulative data rows for the REPORT totals.
# "JED - Obhor" covers two Obhor sheets combined; "JED - JDR" has no sheet yet.
_REPORT_BRANCH_TO_SHEETS: dict[str, list[str]] = {}
for _sheet, (_gym, _branch) in PAYMENT_SHEET_TO_BRANCH.items():
    _REPORT_BRANCH_TO_SHEETS.setdefault(_branch, []).append(_sheet)

# Special cases in the REPORT sheet that differ from the sheet branch keys:
_REPORT_BRANCH_TO_SHEETS["JED - Obhor"] = [
    "Masters -JED- Obhor - Al Amwaj",
    "Masters -JED- Obhor - Al Sheraa",
]
_REPORT_BRANCH_TO_SHEETS["JED - JDR"] = []   # no sheet in template (yet)


def _copy_row_style(src_ws, src_row: int, dst_ws, dst_row: int, max_col: int):
    for col in range(1, max_col + 1):
        src = src_ws.cell(row=src_row, column=col)
        dst = dst_ws.cell(row=dst_row, column=col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def _last_data_row(ws) -> int:
    """Return the last row (1-based) that has a value in col E (dispatch_date)."""
    last = 6  # default: right before row 7
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=5, max_col=5):
        if row[0].value is not None:
            last = row[0].row
    return last


def _count_data_rows(ws) -> int:
    """Count rows from row 7 down that have a value in col E."""
    count = 0
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=5, max_col=5):
        if row[0].value is not None:
            count += 1
    return count


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def generate_payment_report(
    programs: list[dict],   # from Supabase (both gyms), keys: gym, branch, client_id,
                            # client_name, trainer_name, test_date, dispatch_date
    month: int,
    year: int,
    report_date: date | None = None,
) -> bytes:
    """
    Append a month's programs to the cumulative payment template and return
    the result as bytes. The original template file is never modified.
    """
    if not os.path.exists(PAYMENT_TEMPLATE_PATH):
        raise FileNotFoundError(f"Payment template not found: {PAYMENT_TEMPLATE_PATH}")

    with open(PAYMENT_TEMPLATE_PATH, "rb") as f:
        payment_file_bytes = f.read()
    import calendar as _cal
    month_start = date(year, month, 1)
    month_end = date(year, month, _cal.monthrange(year, month)[1])

    def in_month(p) -> bool:
        dd = _to_date(p.get("dispatch_date"))
        return dd is not None and month_start <= dd <= month_end

    monthly = [p for p in programs if in_month(p)]

    # Group by (gym, branch)
    by_gym_branch: dict[tuple[str, str], list[dict]] = {}
    for p in monthly:
        key = (p.get("gym", ""), p.get("branch", ""))
        by_gym_branch.setdefault(key, []).append(p)

    # Sort each branch's programs by dispatch_date then test_date
    for key in by_gym_branch:
        by_gym_branch[key].sort(key=lambda p: (
            _to_date(p.get("dispatch_date")) or date.min,
            _to_date(p.get("test_date")) or date.min,
        ))

    wb = load_workbook(io.BytesIO(payment_file_bytes), data_only=False)  # template copy, never overwrites
    rpt_date = report_date or date.today()
    date_fmt = '[$-1010000]d/m/yyyy;@'

    for sheet_name in wb.sheetnames:
        if sheet_name == "REPORT":
            continue

        mapping = PAYMENT_SHEET_TO_BRANCH.get(sheet_name)
        if mapping is None:
            continue  # unknown sheet, leave untouched

        gym, branch = mapping
        new_rows = by_gym_branch.get((gym, branch), [])

        ws = wb[sheet_name]

        # Always insert a green separator (even if no new rows — acts as a month marker)
        last_row = _last_data_row(ws)
        sep_row = last_row + 1

        # Write green separator row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=sep_row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = copy(GREEN_FILL)

        if not new_rows:
            continue

        # Style reference: use last existing data row for formatting
        style_ref = last_row if last_row >= 7 else 7

        for i, prog in enumerate(new_rows):
            dest_row = sep_row + 1 + i
            _copy_row_style(ws, style_ref, ws, dest_row, ws.max_column)

            client_id     = prog.get("client_id") or None
            client_name   = prog.get("client_name", "")
            trainer_name  = prog.get("trainer_name", "") or ""
            test_date     = _to_date(prog.get("test_date"))
            dispatch_date = _to_date(prog.get("dispatch_date"))

            def _set(col, val, fmt=None):
                c = ws.cell(row=dest_row, column=col)
                if isinstance(c, MergedCell):
                    return
                c.value = val
                if fmt:
                    c.number_format = fmt

            _set(1, client_id)
            _set(2, client_name)
            _set(3, trainer_name)
            _set(4, test_date, date_fmt)
            _set(5, dispatch_date, date_fmt)

    # ── Update REPORT sheet ───────────────────────────────────────────────────
    rpt_ws = wb["REPORT"]
    rpt_ws["B3"] = rpt_date
    rpt_ws["B3"].number_format = date_fmt

    # Build sheet→count map (after all appends)
    sheet_counts: dict[str, int] = {
        sn: _count_data_rows(wb[sn])
        for sn in wb.sheetnames
        if sn != "REPORT"
    }

    def _total_for_report_branch(branch_label: str) -> int:
        sheets = _REPORT_BRANCH_TO_SHEETS.get(branch_label, [])
        return sum(sheet_counts.get(s, 0) for s in sheets)

    # Fill Masters totals (col B, starting row 9) and Motions totals (col E, starting row 9)
    row = 9
    while True:
        masters_label = rpt_ws.cell(row=row, column=1).value
        motions_label = rpt_ws.cell(row=row, column=4).value
        if masters_label is None and motions_label is None:
            break

        if masters_label:
            rpt_ws.cell(row=row, column=2).value = _total_for_report_branch(str(masters_label).strip())
        if motions_label:
            rpt_ws.cell(row=row, column=5).value = _total_for_report_branch(str(motions_label).strip())

        row += 1
        if row > 200:
            break  # safety guard

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
