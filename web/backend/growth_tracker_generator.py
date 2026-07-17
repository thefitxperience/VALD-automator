"""
Test Growth Tracker report generator.

Month-over-month comparison of test counts (previous month vs selected month),
counted by dispatch_date over approved, non-ignored programs.

Layout mirrors the manual "Test Growth Tracker" workbooks:
  • "Branches analysis" sheet — per-branch counts (prev, curr), Difference, Growth Rate %,
    Total row, plus side tables: "Branches with 0 tests" and "Top 3 Branches" (per month).
  • Detail sheets:
      - Body Motions: one sheet per branch  → per-trainer breakdown.
      - Body Masters: one sheet per area manager → grouped by branch → per-trainer breakdown,
        with the manager's own Top-3 / 0-tests header.
"""
import io
import calendar
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
from openpyxl.utils import get_column_letter

from report_generator import BRANCH_ORDER

# Programs whose branch/trainer is blank use these sentinels; they are excluded from the report.
_NO_BRANCH = "(no branch)"
_NO_TRAINER = "(no trainer)"

# ── Body Masters area-manager → branches (hardcoded, normalized to DB branch names) ──
# Order = sheet order in the workbook. Branch names must match the `programs.branch` values.
MANAGER_BRANCHES: list[tuple[str, list[str]]] = [
    ("Mohamed Tantawi", [
        "RUH - Ezdehar", "RUH - Eshbilia", "RUH - Rabwa", "RUH - Al Kharj",
        "RUH - Al Aarid", "Hail",
    ]),
    ("Khalil Hammami", [
        "RUH - Salam", "RUH - Al Wadi", "RUH - Al Sahafa", "ALQ - Buraidah",
        "ALQ - Unaizah", "ALQ - Al Rass", "ALQ - Al Rayyan", "AlUla",
    ]),
    ("Karim Husni", [
        "RUH - Murooj", "RUH - Al Nahda", "RUH - Swaidi", "RUH - Al Badia",
        "RUH - Badr", "RUH - Takhasousi", "Tabuk",
    ]),
    ("Zakaria Maghmouma", [
        "RUH - Al Khaleej", "RUH - Shubra", "RUH - Muzahmiyah", "RUH - Al Malaz",
        "RUH - Al Fayha", "RUH - Al Massif", "Najran",
    ]),
    ("Mohammaed Reda", [
        "DMM - Al Athir", "DMM - Al Jameyeen", "DMM - Khobar", "Al Mubaraz",
        "DMM - Hufof", "Hafr El Batin", "Uhud",
    ]),
    ("Mulfi Ahmed Al-Ghamdi", [
        "MED - Taiba", "MED - Shouran", "JED - Obhor - Al Amwaj", "JED - Obhor - Al Sheraa",
        "JED - Hamadania", "JED - Makkah", "Khamis Mushait", "JED - Al Rawdah",
    ]),
]

# ── Styling (matches the manual tracker workbooks) ───────────────────────────────
# Header/total fill = theme accent 3 (blue #5B9BD5) at tint 0.8 → light blue.
_HEADER_FILL = PatternFill("solid", fgColor=Color(theme=3, tint=0.8))
_SUBHEAD_FILL = _HEADER_FILL
_TOTAL_FILL = _HEADER_FILL
_BRANCH_FILL = PatternFill("solid", fgColor="FFFFF3B4")   # light gold for the branch-name row

_MONTH_COL_WIDTH = 15
_FONT = Font(name="Calibri", size=11)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True)   # black text on light-blue fill
_BOLD = _HEADER_FONT
_FONT_RED = Font(name="Calibri", size=11, color="FFFF0000")
_BOLD_RED = Font(name="Calibri", size=11, bold=True, color="FFFF0000")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_MONTH_FMT = "mmm\\-yy"      # e.g. Apr-26
_PCT_FMT = "0%"
_thin = Side(style="thin", color="808080")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _prev_month(year: int, month: int) -> tuple[int, int]:
    return (year - 1, 12) if month == 1 else (year, month - 1)


def _ym(d) -> str:
    return (d or "")[:7] if isinstance(d, str) else (d.isoformat()[:7] if d else "")


def _growth(prev: int, curr: int):
    """Growth rate cell value following the manual tracker's rules."""
    if prev == 0 and curr == 0:
        return 0
    if prev == 0 and curr > 0:
        return "New"
    return (curr - prev) / prev


def _mlabel(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]} {year}"


def _style_cell(c, *, fill=None, font=None, align=None, border=True, fmt=None):
    if fill:
        c.fill = fill
    c.font = font or _FONT
    c.alignment = align or _CENTER
    if border:
        c.border = _BORDER
    if fmt:
        c.number_format = fmt
    return c


class _Counter:
    """Counts approved, non-ignored programs by (branch, trainer) for two months (by dispatch_date)."""

    def __init__(self, programs, prev_ym: str, curr_ym: str):
        # branch -> trainer -> [prev, curr]
        self.data: dict[str, dict[str, list[int]]] = {}
        for p in programs:
            if not p.get("approved") or p.get("ignored"):
                continue
            m = _ym(p.get("dispatch_date"))
            if m == prev_ym:
                idx = 0
            elif m == curr_ym:
                idx = 1
            else:
                continue
            branch = p.get("branch") or "(no branch)"
            trainer = p.get("trainer_name") or "(no trainer)"
            self.data.setdefault(branch, {}).setdefault(trainer, [0, 0])[idx] += 1

    def branch_totals(self, branch: str) -> tuple[int, int]:
        prev = curr = 0
        for t in self.data.get(branch, {}).values():
            prev += t[0]
            curr += t[1]
        return prev, curr

    def trainers_for(self, branch: str, roster: list[str]) -> list[str]:
        """Roster order first (from trainers table), then any extra trainers found in programs."""
        found = self.data.get(branch, {})
        ordered = [t for t in roster if t in found or True]  # keep full roster
        extras = sorted(t for t in found if t not in roster)
        # Avoid duplicates while preserving order
        seen, out = set(), []
        for t in list(roster) + extras:
            if t not in seen:
                seen.add(t)
                out.append(t)
        return out

    def count(self, branch: str, trainer: str) -> list[int]:
        return self.data.get(branch, {}).get(trainer, [0, 0])


def _write_count_header(ws, row: int, prev_date, curr_date, first_col_title: str):
    """Writes the 2-row 'Trainer/Branch | tests(prev,curr) | Difference | Growth Rate' header."""
    ws.cell(row, 1, first_col_title)
    ws.cell(row, 2, "Number of tests per Month")
    ws.cell(row, 4, "Difference")
    ws.cell(row, 5, "Growth Rate")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    ws.merge_cells(start_row=row, start_column=4, end_row=row + 1, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row + 1, end_column=5)
    ws.cell(row + 1, 2, prev_date)
    ws.cell(row + 1, 3, curr_date)
    # Style EVERY cell in both header rows (incl. the hidden cells of each merge) so borders
    # are complete on all sides of the merged "Difference" / "Growth Rate" / count headers.
    for col in (1, 2, 3, 4, 5):
        _style_cell(ws.cell(row, col), fill=_HEADER_FILL, font=_HEADER_FONT)
    _style_cell(ws.cell(row + 1, 1), fill=_HEADER_FILL, font=_HEADER_FONT)
    _style_cell(ws.cell(row + 1, 2), fill=_HEADER_FILL, font=_HEADER_FONT, fmt=_MONTH_FMT)
    _style_cell(ws.cell(row + 1, 3), fill=_HEADER_FILL, font=_HEADER_FONT, fmt=_MONTH_FMT)
    _style_cell(ws.cell(row + 1, 4), fill=_HEADER_FILL, font=_HEADER_FONT)
    _style_cell(ws.cell(row + 1, 5), fill=_HEADER_FILL, font=_HEADER_FONT)
    return row + 2


def _write_trainer_rows(ws, start_row: int, branch: str, counter: _Counter, roster: list[str]) -> int:
    r = start_row
    for trainer in counter.trainers_for(branch, roster):
        prev, curr = counter.count(branch, trainer)
        _style_cell(ws.cell(r, 1, trainer), align=_LEFT)
        _style_cell(ws.cell(r, 2, prev))
        _style_cell(ws.cell(r, 3, curr))
        _style_cell(ws.cell(r, 4, curr - prev))
        g = _growth(prev, curr)
        _style_cell(ws.cell(r, 5, g), fmt=("0%" if isinstance(g, (int, float)) and not isinstance(g, bool) else None))
        r += 1
    # Total row
    prev, curr = counter.branch_totals(branch)
    _style_cell(ws.cell(r, 1, "Total tests"), font=_BOLD, fill=_TOTAL_FILL, align=_LEFT)
    _style_cell(ws.cell(r, 2, prev), font=_BOLD, fill=_TOTAL_FILL)
    _style_cell(ws.cell(r, 3, curr), font=_BOLD, fill=_TOTAL_FILL)
    _style_cell(ws.cell(r, 4, curr - prev), font=_BOLD, fill=_TOTAL_FILL)
    g = _growth(prev, curr)
    _style_cell(ws.cell(r, 5, g), font=_BOLD, fill=_TOTAL_FILL,
                fmt=("0%" if isinstance(g, (int, float)) and not isinstance(g, bool) else None))
    return r + 1, f"E{start_row}:E{r}"


def _set_widths(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _redden_zeros(wb):
    """Make any numeric 0 in the count/difference columns (B, C, D) red. Growth Rate (E) is left alone."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=2, max_col=4):
            for c in row:
                v = c.value
                if isinstance(v, (int, float)) and not isinstance(v, bool) and v == 0:
                    c.font = _BOLD_RED if (c.font and c.font.bold) else _FONT_RED


def _add_growth_arrows(ws, ranges: list[str]):
    """Add the 3-arrows icon set to Growth Rate cells: <0 down (red), 0–10% flat, ≥10% up (green).
    Mirrors the manual tracker's cfvo thresholds [percent 0, num 0, num 0.1]."""
    ranges = [r for r in ranges if r]
    if not ranges:
        return
    icon_set = IconSet(
        iconSet="3Arrows",
        cfvo=[
            FormatObject(type="percent", val=0),
            FormatObject(type="num", val=0),
            FormatObject(type="num", val=0.1),
        ],
        showValue=True,
    )
    rule = Rule(type="iconSet", iconSet=icon_set)
    ws.conditional_formatting.add(" ".join(ranges), rule)


def _branch_analysis_sheet(wb, gym, branches, counter, prev_date, curr_date):
    ws = wb.create_sheet("Branches analysis", 0)
    # Side-table + top-3 data
    zero_prev = [b for b in branches if counter.branch_totals(b)[0] == 0]
    zero_curr = [b for b in branches if counter.branch_totals(b)[1] == 0]
    top_prev = [b for b, _ in sorted(branches_with_totals(branches, counter, 0), key=lambda x: -x[1])[:3]]
    top_curr = [b for b, _ in sorted(branches_with_totals(branches, counter, 1), key=lambda x: -x[1])[:3]]

    start = _write_count_header(ws, 1, prev_date, curr_date, "Branch name")
    # Side table headers (cols G,H = 0-tests; J,K = top 3)
    _style_cell(ws.cell(1, 7, "Name of Branches with 0 tests"), fill=_HEADER_FILL, font=_HEADER_FONT)
    _style_cell(ws.cell(1, 8), fill=_HEADER_FILL, font=_HEADER_FONT)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    _style_cell(ws.cell(1, 10, "Top 3 Branches of the Month"), fill=_HEADER_FILL, font=_HEADER_FONT)
    _style_cell(ws.cell(1, 11), fill=_HEADER_FILL, font=_HEADER_FONT)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=11)
    for col, dt in ((7, prev_date), (8, curr_date), (10, prev_date), (11, curr_date)):
        _style_cell(ws.cell(2, col, dt), fill=_SUBHEAD_FILL, font=_BOLD, fmt=_MONTH_FMT)

    r = start
    for b in branches:
        prev, curr = counter.branch_totals(b)
        _style_cell(ws.cell(r, 1, b), align=_LEFT)
        _style_cell(ws.cell(r, 2, prev))
        _style_cell(ws.cell(r, 3, curr))
        _style_cell(ws.cell(r, 4, curr - prev))
        g = _growth(prev, curr)
        _style_cell(ws.cell(r, 5, g), fmt=("0%" if isinstance(g, (int, float)) and not isinstance(g, bool) else None))
        r += 1
    # Total
    tp = sum(counter.branch_totals(b)[0] for b in branches)
    tc = sum(counter.branch_totals(b)[1] for b in branches)
    _style_cell(ws.cell(r, 1, "Total"), font=_BOLD, fill=_TOTAL_FILL, align=_LEFT)
    _style_cell(ws.cell(r, 2, tp), font=_BOLD, fill=_TOTAL_FILL)
    _style_cell(ws.cell(r, 3, tc), font=_BOLD, fill=_TOTAL_FILL)
    _style_cell(ws.cell(r, 4, tc - tp), font=_BOLD, fill=_TOTAL_FILL)
    g = _growth(tp, tc)
    _style_cell(ws.cell(r, 5, g), font=_BOLD, fill=_TOTAL_FILL,
                fmt=("0%" if isinstance(g, (int, float)) and not isinstance(g, bool) else None))
    _add_growth_arrows(ws, [f"E{start}:E{r}"])

    # Fill side tables — always border both columns down to the taller of the two months,
    # so a shorter month's empty cells still complete the box instead of leaving it ragged.
    for i in range(max(len(zero_prev), len(zero_curr))):
        _style_cell(ws.cell(3 + i, 7, zero_prev[i] if i < len(zero_prev) else None), align=_LEFT)
        _style_cell(ws.cell(3 + i, 8, zero_curr[i] if i < len(zero_curr) else None), align=_LEFT)
    for i in range(max(len(top_prev), len(top_curr))):
        _style_cell(ws.cell(3 + i, 10, top_prev[i] if i < len(top_prev) else None), align=_LEFT)
        _style_cell(ws.cell(3 + i, 11, top_curr[i] if i < len(top_curr) else None), align=_LEFT)

    _set_widths(ws, {1: 22, 2: 11, 3: 11, 4: 11, 5: 12, 6: 3, 7: 20, 8: 20, 9: 3, 10: 20, 11: 20})
    return ws


def branches_with_totals(branches, counter, idx):
    return [(b, counter.branch_totals(b)[idx]) for b in branches]


def generate_growth_tracker(
    gym: str,
    programs: list[dict],
    year: int,
    month: int,
    trainer_order_by_branch: dict | None = None,
    report_date: date | None = None,
) -> bytes:
    prev_y, prev_m = _prev_month(year, month)
    prev_ym = f"{prev_y}-{prev_m:02d}"
    curr_ym = f"{year}-{month:02d}"
    prev_date = date(prev_y, prev_m, 1)
    curr_date = date(year, month, 1)

    counter = _Counter(programs, prev_ym, curr_ym)
    roster_by_branch = trainer_order_by_branch or {}
    branches = list(BRANCH_ORDER.get(gym, []))
    # Include any branch that has data but isn't in BRANCH_ORDER (keeps totals honest),
    # but skip the "(no branch)" sentinel — unbranded programs are excluded from the report.
    for b in counter.data:
        if b not in branches and b != _NO_BRANCH:
            branches.append(b)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _branch_analysis_sheet(wb, gym, branches, counter, prev_date, curr_date)

    if gym == "Body Motions":
        # One sheet per branch
        for b in branches:
            ws = wb.create_sheet(_safe_title(b))
            start = _write_count_header(ws, 1, prev_date, curr_date, "Trainer name")
            _, grange = _write_trainer_rows(ws, start, b, counter, roster_by_branch.get(b, []))
            _add_growth_arrows(ws, [grange])
            _set_widths(ws, {1: 28, 2: 11, 3: 11, 4: 11, 5: 12})
    else:
        # Body Masters: one sheet per manager, grouped by branch
        assigned = {b for _, bs in MANAGER_BRANCHES for b in bs}
        for manager, mbranches in MANAGER_BRANCHES:
            ws = wb.create_sheet(_safe_title(manager))
            # manager header: top-3 / 0-tests for this manager's branches
            mtop_prev = [b for b, _ in sorted(branches_with_totals(mbranches, counter, 0), key=lambda x: -x[1])[:3]]
            mtop_curr = [b for b, _ in sorted(branches_with_totals(mbranches, counter, 1), key=lambda x: -x[1])[:3]]
            mzero_prev = [b for b in mbranches if counter.branch_totals(b)[0] == 0]
            mzero_curr = [b for b in mbranches if counter.branch_totals(b)[1] == 0]
            _style_cell(ws.cell(1, 1, "Top 3 Branches of the Month"), fill=_HEADER_FILL, font=_HEADER_FONT)
            _style_cell(ws.cell(1, 2), fill=_HEADER_FILL, font=_HEADER_FONT)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            _style_cell(ws.cell(1, 4, "Name of Branches with 0 tests"), fill=_HEADER_FILL, font=_HEADER_FONT)
            _style_cell(ws.cell(1, 5), fill=_HEADER_FILL, font=_HEADER_FONT)
            ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
            for col, dt in ((1, prev_date), (2, curr_date), (4, prev_date), (5, curr_date)):
                _style_cell(ws.cell(2, col, dt), fill=_SUBHEAD_FILL, font=_BOLD, fmt=_MONTH_FMT)
            # Border both columns down to the taller month so each side table stays rectangular.
            for i in range(max(len(mtop_prev), len(mtop_curr))):
                _style_cell(ws.cell(3 + i, 1, mtop_prev[i] if i < len(mtop_prev) else None), align=_LEFT)
                _style_cell(ws.cell(3 + i, 2, mtop_curr[i] if i < len(mtop_curr) else None), align=_LEFT)
            for i in range(max(len(mzero_prev), len(mzero_curr))):
                _style_cell(ws.cell(3 + i, 4, mzero_prev[i] if i < len(mzero_prev) else None), align=_LEFT)
                _style_cell(ws.cell(3 + i, 5, mzero_curr[i] if i < len(mzero_curr) else None), align=_LEFT)

            # Side tables start at row 3; start the branch tables one blank row after the tallest one.
            last_side_row = 2 + max(len(mtop_prev), len(mtop_curr), len(mzero_prev), len(mzero_curr))
            r = last_side_row + 2
            granges = []
            for b in mbranches:
                # "Branch name" label (col A, centered) + branch name merged & centered across B:E
                _style_cell(ws.cell(r, 1, "Branch name"), font=_BOLD, fill=_BRANCH_FILL)
                ws.cell(r, 2, b)
                for col in (2, 3, 4, 5):
                    _style_cell(ws.cell(r, col), font=_BOLD, fill=_BRANCH_FILL)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
                r += 1
                r = _write_count_header(ws, r, prev_date, curr_date, "Trainer name")
                r, grange = _write_trainer_rows(ws, r, b, counter, roster_by_branch.get(b, []))
                granges.append(grange)
                r += 1  # blank spacer
            _add_growth_arrows(ws, granges)
            _set_widths(ws, {1: 28, 2: _MONTH_COL_WIDTH, 3: _MONTH_COL_WIDTH,
                             4: _MONTH_COL_WIDTH, 5: _MONTH_COL_WIDTH})

        # Branches with no manager assigned (e.g. a newly added branch) → note sheet
        unassigned = [b for b in branches if b not in assigned]
        if unassigned:
            ws = wb.create_sheet("Unassigned branches")
            _style_cell(ws.cell(1, 1, "Branches not assigned to any manager"), font=_BOLD, align=_LEFT, border=False)
            start = _write_count_header(ws, 2, prev_date, curr_date, "Branch name")
            r = start
            for b in unassigned:
                prev, curr = counter.branch_totals(b)
                _style_cell(ws.cell(r, 1, b), align=_LEFT)
                _style_cell(ws.cell(r, 2, prev))
                _style_cell(ws.cell(r, 3, curr))
                _style_cell(ws.cell(r, 4, curr - prev))
                g = _growth(prev, curr)
                _style_cell(ws.cell(r, 5, g), fmt=("0%" if isinstance(g, (int, float)) and not isinstance(g, bool) else None))
                r += 1
            _add_growth_arrows(ws, [f"E{start}:E{r - 1}"])
            _set_widths(ws, {1: 24, 2: 11, 3: 11, 4: 11, 5: 12})

    _redden_zeros(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _safe_title(name: str) -> str:
    """Excel sheet titles: max 31 chars, no []:*?/\\ ."""
    t = name.strip()
    for ch in '[]:*?/\\':
        t = t.replace(ch, " ")
    return t[:31]
