"""
Bodydot monthly report generator.

Fills the "Bodydot Month YEAR - <gym>.xlsx" templates, which mirror the VALD
monthly report (REPORT / REPORT 2 / per-branch data sheet, dispatch-date driven)
plus a TEST VALIDITY summary sheet.

Two data sources, matching the agreed design:
  • TEST VALIDITY (Total / Valid / Invalid) — LIVE from the Bodydot API for the
    month, by test date (invalid tests never get approved, so they can't come
    from our DB).
  • REPORT / REPORT 2 / data sheet — from APPROVED rows in `bodydot_tests`,
    dispatch-date driven, exactly like the VALD report. Per-trainer counts are
    Excel COUNTIF formulas over the data sheet's TRAINER NAME column.

Bodydot has one branch per gym, so there is a single data sheet:
    Body Masters → "RUH - Al Aarid"   Body Motions → "RUH - Al Sahafa"
"""
import io
import os
import re
from copy import copy
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.cell.cell import MergedCell  # merged (non-anchor) cells are read-only

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_MAP = {
    "Body Masters": os.path.join(BASE_DIR, "Bodydot Month YEAR - Body Masters.xlsx"),
    "Body Motions": os.path.join(BASE_DIR, "Bodydot Month YEAR - Body Motions.xlsx"),
}

_SUMMARY_SHEETS = ("TEST VALIDITY", "REPORT", "REPORT 2")


def _as_date(v):
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def _in_window(d, start: date, end: date):
    d = _as_date(d)
    return bool(d) and start <= d <= end


def _client_ref(row) -> str:
    """The report's CLIENT ID column. Bodydot's own id isn't the gym's id, so prefer
    the manually-set `real_client_id`; otherwise fall back to a number embedded in the
    client name (trainers often type the id into the name), else blank."""
    rid = (row.get("real_client_id") or "").strip()
    if rid:
        return rid
    nums = re.findall(r"\d{4,}", row.get("client_name") or "")
    return max(nums, key=len) if nums else ""


def _client_display_name(row) -> str:
    """The report's CLIENT NAME column, with any embedded id number stripped out
    (the id lives in its own column) — e.g. 'Sabrina 12345' → 'Sabrina'."""
    name = row.get("client_name") or ""
    cleaned = re.sub(r"\s*\d{4,}\s*", " ", name).strip()
    return cleaned or name


def _rebuild_report_sheet(ws, branch, trainers):
    """Rewrite the REPORT sheet's trainer rows (single branch) with COUNTIF /
    COUNTIFS formulas, copying styling from template rows 7 (branch) and 8 (cont)."""
    max_col = ws.max_column
    row_height = ws.row_dimensions[7].height or 50

    def capture(row):
        return {c: {
            "font": copy(ws.cell(row, c).font),
            "fill": copy(ws.cell(row, c).fill),
            "alignment": copy(ws.cell(row, c).alignment),
            "border": copy(ws.cell(row, c).border),
            "number_format": ws.cell(row, c).number_format,
        } for c in range(1, max_col + 1)}

    branch_styles = capture(7)
    cont_styles = capture(8)

    for r in range(7, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, trainer in enumerate(trainers):
        r = 7 + i
        styles = branch_styles if i == 0 else cont_styles
        for c in range(1, max_col + 1):
            s = styles[c]
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.font = copy(s["font"])
            cell.fill = copy(s["fill"])
            cell.alignment = copy(s["alignment"])
            cell.border = copy(s["border"])
            cell.number_format = s["number_format"]
        if i == 0:
            ws.cell(r, 1).value = branch
        ws.cell(r, 2).value = trainer
        ws.cell(r, 3).value = (
            f'=IFERROR(\n   COUNTIF(INDIRECT("\'" & LOOKUP("zzz",$A$7:A{r}) & "\'!C:C"), B{r}),\n   0\n)'
        )
        d_formula = (
            f'=IFERROR(\n  COUNTIFS(\n'
            f'    INDIRECT("\'" & LOOKUP(2,1/($A$7:A{r}<>""),$A$7:A{r}) & "\'!C:C"), $B{r},\n'
            f'    INDIRECT("\'" & LOOKUP(2,1/($A$7:A{r}<>""),$A$7:A{r}) & "\'!E:E"), ">=" & ($B$3-6),\n'
            f'    INDIRECT("\'" & LOOKUP(2,1/($A$7:A{r}<>""),$A$7:A{r}) & "\'!E:E"), "<=" & $B$3\n'
            f'  ),\n0)'
        )
        ws.cell(r, 4).value = ArrayFormula(f"D{r}", d_formula)
        ws.row_dimensions[r].height = row_height


def _copy_row_style(ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        src, dst = ws.cell(src_row, c), ws.cell(dst_row, c)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def generate_bodydot_report(
    gym: str,
    period_start: date,
    period_end: date,
    validity: dict,            # {"total": int, "valid": int, "invalid": int}
    approved_tests: list[dict],  # bodydot_tests rows (approved)
    trainer_roster: list[str] | None = None,
    report_date: date | None = None,
) -> bytes:
    template_path = TEMPLATE_MAP.get(gym)
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"Bodydot template not found: {template_path}")

    wb = load_workbook(template_path, data_only=False)
    data_sheet_name = next(s for s in wb.sheetnames if s not in _SUMMARY_SHEETS)
    rpt_date = report_date or period_end

    # Programs dispatched within the period window drive the REPORT sheets + data sheet.
    dispatched = [t for t in approved_tests if _in_window(t.get("dispatch_date"), period_start, period_end)]

    # REPORT sheet lists only regular roster trainers — custom/ad-hoc names entered at
    # approval time are excluded (their tests still count in the data sheet + club totals).
    trainers = sorted(set(trainer_roster or []), key=str.lower)

    # ── TEST VALIDITY (live totals) ──
    total = int(validity.get("total", 0))
    valid = int(validity.get("valid", 0))
    invalid = int(validity.get("invalid", 0))
    ws = wb["TEST VALIDITY"]
    ws["B7"], ws["C7"] = total, 1 if total else 0
    ws["B8"], ws["C8"] = valid, (valid / total if total else 0)
    ws["B9"], ws["C9"] = invalid, (invalid / total if total else 0)

    # ── REPORT (per trainer) ──
    ws = wb["REPORT"]
    ws["B3"] = rpt_date
    if trainers:
        _rebuild_report_sheet(ws, data_sheet_name, trainers)

    # ── REPORT 2 (club totals) — formulas already reference the data sheet ──
    wb["REPORT 2"]["B3"] = rpt_date

    # ── Data sheet ──
    ws = wb[data_sheet_name]
    ws["B3"] = rpt_date
    for r in range(7, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for i, t in enumerate(dispatched):
        r = 7 + i
        _copy_row_style(ws, 7, r, ws.max_column)
        ws.cell(r, 1, _client_ref(t))
        ws.cell(r, 2, _client_display_name(t))
        ws.cell(r, 3, t.get("trainer_name") or "")
        td, dd = _as_date(t.get("test_date")), _as_date(t.get("dispatch_date"))
        ws.cell(r, 4, td)
        ws.cell(r, 5, dd)
        ws.cell(r, 4).number_format = "DD/MM/YYYY"
        ws.cell(r, 5).number_format = "DD/MM/YYYY"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
