"""
Server-side check file processor.
Extracted from process_dynamo.py — pure openpyxl, no xlwings.
"""
import re
from datetime import datetime


# ── fast xlsx loading via calamine (Rust) ────────────────────────────────────

def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel column letter(s): 1→A, 27→AA."""
    result = ''
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _make_ws_adapter(rows_data: list):
    """Build a minimal openpyxl-compatible worksheet from calamine row data."""
    cells = {}
    for r_idx, row in enumerate(rows_data):
        row_num = r_idx + 1
        for c_idx, val in enumerate(row):
            if val != '':   # calamine uses '' for empty, openpyxl uses None
                cells[f"{_col_letter(c_idx + 1)}{row_num}"] = val
    _max_row = len(rows_data)

    class _Cell:
        __slots__ = ('value',)
        def __init__(self, v): self.value = v

    class _Ws:
        def __getitem__(self, key): return _Cell(cells.get(key))
        @property
        def max_row(self): return _max_row

    return _Ws()


def _load_worksheet(file_bytes: bytes):
    """Load the first sheet. Uses calamine (fast) if available, openpyxl otherwise."""
    try:
        from python_calamine import CalamineWorkbook
        import io
        cal = CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
        ws = _make_ws_adapter(cal.get_sheet_by_index(0).to_python())
        cal.close()
        return ws, None          # (worksheet, workbook_or_None)
    except ImportError:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        return wb.active, wb


# ── helpers ──────────────────────────────────────────────────────────────────

def nz_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def nz_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def normalize_test_date(test_date):
    if test_date is None:
        return None
    if isinstance(test_date, datetime):
        return test_date.strftime("%Y-%m-%d")
    s = str(test_date).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Try ISO prefix (e.g. "2026-04-01 00:00:00")
    if len(s) >= 10:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s


def parse_asymmetry(raw):
    if raw is None:
        return None, None
    s = str(raw).strip().lower()
    if not s:
        return None, None
    if s in ("n/a", "na", "n.a.", "n.a"):
        return None, None
    if "%" not in s:
        try:
            return float(s.replace(",", ".")), None
        except ValueError:
            return None, None
    percent_index = s.index("%")
    num_part = s[:percent_index].strip()
    if num_part.lower() in ("n/a", "na", "n.a.", "n.a"):
        return None, None
    side_char = None
    for ch in reversed(s):
        if ch != " ":
            side_char = ch
            break
    try:
        return float(num_part.replace(",", ".")), side_char
    except ValueError:
        return None, None


def get_movement_test_type(movement, region):
    m = movement.lower().strip()
    r = region.lower().strip()
    upper = [
        r == "shoulder" and m in ("external rotation", "internal rotation", "flexion", "abduction", "push", "pull"),
        r == "hand",
        r == "elbow" and m in ("extension", "flexion"),
    ]
    lower = [
        r == "trunk",
        r == "knee",
        r == "hip" and m in ("flexion", "extension", "abduction", "adduction"),
    ]
    if any(upper):
        return "upper"
    if any(lower):
        return "lower"
    return None


def detect_test_type(patient_rows, src_ws):
    movements_regions = set()
    for row in patient_rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region = nz_str(src_ws[f"H{row}"].value).lower().strip()
        if movement and region:
            movements_regions.add((movement, region))

    upper_indicators = {
        ("external rotation", "shoulder"), ("internal rotation", "shoulder"),
        ("flexion", "shoulder"), ("abduction", "shoulder"),
        ("push", "shoulder"), ("pull", "shoulder"), ("grip squeeze", "hand"),
    }
    elbow_indicators = {("extension", "elbow"), ("flexion", "elbow")}
    lower_indicators = {("lateral flexion", "trunk"), ("flexion", "hip"), ("extension", "hip")}
    knee_indicators = {("extension", "knee"), ("flexion", "knee")}
    hip_abd_add_indicators = {("abduction", "hip"), ("adduction", "hip")}

    has_upper = any(mr in movements_regions for mr in upper_indicators)
    has_elbow = any(mr in movements_regions for mr in elbow_indicators)
    has_knee = any(mr in movements_regions for mr in knee_indicators)
    has_hip_abd_add = any(mr in movements_regions for mr in hip_abd_add_indicators)
    has_lower_specific = any(mr in movements_regions for mr in lower_indicators)
    has_trunk = ("lateral flexion", "trunk") in movements_regions
    has_hip_flex = ("flexion", "hip") in movements_regions

    # Full body: shoulder + elbow + knee, no hip flex/ext or trunk
    if has_upper and has_elbow and has_knee and not has_lower_specific:
        return "full"

    # Full body: elbow + knee only (no shoulder), no hip flex/ext or trunk
    if has_elbow and has_knee and not has_upper and not has_lower_specific:
        return "full"

    # Full body: elbow + knee + hip_abd/add (hip extension ignored), no trunk or hip flexion
    # e.g. knee ext/flex + hip abd/add + hip ext + elbow ext/flex + shoulder IR/ER
    if has_elbow and has_knee and has_hip_abd_add and not has_trunk and not has_hip_flex:
        return "full"

    has_any_upper = has_upper or has_elbow
    has_any_lower = has_knee or has_hip_abd_add or has_lower_specific

    if has_any_upper and has_any_lower:
        return ["upper", "lower"]
    if has_any_upper:
        return "upper"
    if has_any_lower:
        return "lower"
    return "upper"


def calculate_trunk_asymmetry(patient_rows, src_ws):
    trunk_data = {}
    for row in patient_rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region = nz_str(src_ws[f"H{row}"].value).lower().strip()
        if region == "trunk" and "lateral flexion" in movement:
            force = nz_float(src_ws[f"O{row}"].value)
            if "right" in movement:
                trunk_data["right"] = force
            elif "left" in movement:
                trunk_data["left"] = force
    if "right" not in trunk_data or "left" not in trunk_data:
        return None, None
    r, l = trunk_data["right"], trunk_data["left"]
    if r <= 0 or l <= 0:
        return None, None
    avg = (l + r) / 2
    pct = abs(((l - r) / avg) * 100)
    weak = "Right" if r < l else "Left"
    return pct, weak


# ── cell mapping helpers ─────────────────────────────────────────────────────

def get_movement_label(movement, region):
    m, r = movement.lower().strip(), region.lower().strip()
    labels = {
        ('internal rotation', 'shoulder'): "Shoulder IR Standing Asymmetry / عدم توازن دوران الكتف الداخلي",
        ('external rotation', 'shoulder'): "Shoulder External Rotation Asymmetry /  عدم توازن الدوران الخارجي للكتف",
        ('flexion', 'shoulder'): "Shoulder Flexion Asymmetry /\n عدم تناسق انثناء الكتف",
        ('abduction', 'shoulder'): "Shoulder Abduction Asymmetry /\nعدم توازن إبعاد الكتف",
        ('push', 'shoulder'): "Shoulder Push Asymmetry /\nعدم توازن في دفع الكتف",
        ('pull', 'shoulder'): "Shoulder Pull Asymmetry /\nعدم توازن في سحب الكتف",
        ('extension', 'elbow'): "Elbow Extension Asymmetry /\nعدم توازن تمديد الكوع",
        ('flexion', 'elbow'): "Elbow Flexion Asymmetry /\nعدم توازن انثناء الكوع",
        ('grip squeeze', 'hand'): "Grip Squeeze Asymmetry /\nعدم توازن ضغط القبضة",
        ('extension', 'knee'): "Knee Extension Asymmetry /\n عدم تناسق تمديد الركبة",
        ('flexion', 'knee'): "Knee Flexion Asymmetry /\n عدم تناسق انثناء الركبة",
        ('abduction', 'hip'): "Hip Abduction Asymmetry /\n عدم تناسق إبعاد الورك",
        ('adduction', 'hip'): "Hip Adduction Asymmetry /\n عدم تناسق تقريب الورك",
        ('lateral flexion', 'trunk'): "Trunk Lateral Flexion /\nالثني الجانبي للجذع",
        ('flexion', 'hip'): "Hip Flexion Asymmetry /\nعدم تناسق ثني الورك",
        ('extension', 'hip'): "Hip Extension Asymmetry /\nعدم تناسق مدّ الورك",
    }
    return labels.get((m, r))


def get_remark_for_percentage(pct_fraction):
    """pct_fraction is 0–1 (as stored in Excel cell formatted as %)."""
    if pct_fraction is None:
        return ""
    pct = abs(pct_fraction) * 100
    if 0.1 <= pct <= 3.9:   return "Perfect Symmetry / \nتناظر مثالي"
    elif 4 <= pct <= 7.9:   return "Normal Symmetry / \nتناظر طبيعي"
    elif 8 <= pct <= 14.9:  return "Weakness / \nضعف"
    elif 15 <= pct <= 19.9: return "Problem / \nمشكلة"
    elif 20 <= pct <= 29.9: return "Major Problem / \nمشكلة كبيرة"
    elif pct >= 30:         return "Risk Of Injury / \nخطر الإصابة"
    return ""


def get_lower_body_cells(m, r, knee_movements, hip_abd_add_movements, hip_flex_ext_movements):
    if r == "knee":
        for mv, right_txt, left_txt in [
            ("extension", "Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى", "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"),
            ("flexion",   "Right Hamstring /\nعضلات الفخذ الخلفية اليمنى",   "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"),
        ]:
            if m == mv and (mv, r) in knee_movements:
                idx = knee_movements.index((mv, r))
                cell = [("C21","D21","C22"), ("C23","D23","C24")][idx] if idx < 2 else None
                if cell: return cell[0], cell[1], cell[2], right_txt, left_txt
    if r == "hip" and m in ["adduction", "abduction"]:
        pairs = [
            ("adduction", "Right Adductors /\nعضلات الفخذ الداخلي اليمنى", "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"),
            ("abduction", "Right Abductors /\nعضلات الفخذ الخارجية اليمنى","Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"),
        ]
        for mv, right_txt, left_txt in pairs:
            if m == mv and (mv, r) in hip_abd_add_movements:
                idx = hip_abd_add_movements.index((mv, r))
                cell = [("O21","P21","O22"), ("O23","P23","O24")][idx] if idx < 2 else None
                if cell: return cell[0], cell[1], cell[2], right_txt, left_txt
    if r == "hip" and m in ["flexion", "extension"]:
        pairs = [
            ("flexion",   "Right Hip Flexors /\nعضلات مثنية الورك اليمنى",   "Left Hip Flexors /\nعضلات مثنية الورك اليسرى"),
            ("extension", "Right Hip Extensors /\nعضلات باسطة الورك اليمنى", "Left Hip Extensors /\nعضلات باسطة الورك اليسرى"),
        ]
        for mv, right_txt, left_txt in pairs:
            if m == mv and (mv, r) in hip_flex_ext_movements:
                idx = hip_flex_ext_movements.index((mv, r))
                cell = [("AG21","AH21","AG22"), ("AG23","AH23","AG24")][idx] if idx < 2 else None
                if cell: return cell[0], cell[1], cell[2], right_txt, left_txt
    return None, None, None, None, None


def get_upper_body_cells(m, r, shoulder_c_movements, shoulder_o_movements, elbow_movements, hand_movements):
    if r == "shoulder" and m in ["external rotation", "internal rotation", "flexion", "abduction"]:
        if (m, r) in shoulder_c_movements:
            idx = shoulder_c_movements.index((m, r))
            slots = [("C21","D21","C22"),("C23","D23","C24"),("C25","D25","C26"),("C27","D27","C28")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "external rotation": ("Right External Rotation /\nالدوران الخارجي الأيمن ", "Left External Rotation /\nالدوران الخارجي الأيسر "),
                    "internal rotation": ("Right Internal Rotation /\n الدوران الداخلي الأيمن", "Left Internal Rotation / \nالدوران الداخلي الأيسر"),
                    "flexion":           ("Right shoulder flexion /\n ثني الكتف الأيمن",        "Left shoulder flexion /\n ثني الكتف الأيسر"),
                    "abduction":         ("Right shoulder abductor /\nعضلة فتح الكتف الأيمن",  "Left shoulder abductor /\nعضلة فتح الكتف الأيسر "),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "shoulder" and m in ["push", "pull"]:
        if (m, r) in shoulder_o_movements:
            idx = shoulder_o_movements.index((m, r))
            slots = [("O21","P21","O22"),("O23","P23","O24")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "push": ("Right Shoulder Push/\nدفع الكتف الأيمن", "Left Shoulder Push/\nدفع الكتف الأيسر"),
                    "pull": ("Right Shoulder Pull/\nسحب الكتف الأيمن", "Left Shoulder Pull/\nسحب الكتف الأيسر"),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "elbow" and m in ["extension", "flexion"]:
        if (m, r) in elbow_movements:
            idx = elbow_movements.index((m, r))
            slots = [("AA21","AB21","AA22"),("AA23","AB23","AA24")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "extension": ("Right Triceps /\n عضلات التراي سيبس اليمنى", "Left Triceps /\n عضلات التراي سيبس اليسرى"),
                    "flexion":   ("Right Biceps /\nعضلة الباي سيبس اليمنى",      "Left Biceps /\n عضلة الباي سيبس اليسرى"),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "hand" and m == "grip squeeze" and (m, r) in hand_movements:
        return "AG21", "AH21", "AG22", "Right Grip Squeeze/\nضغط القبضة باليد اليمنى", "Left Grip Squeeze/\nضغط القبضة باليد اليسرى"
    return None, None, None, None, None


def get_full_body_cells(m, r, shoulder_c_movements, elbow_o_movements, knee_aa_movements, hip_ag_movements):
    if r == "shoulder" and m in ["external rotation", "internal rotation", "flexion", "abduction"]:
        if (m, r) in shoulder_c_movements:
            idx = shoulder_c_movements.index((m, r))
            slots = [("C21","D21","C22"),("C23","D23","C24"),("C25","D25","C26"),("C27","D27","C28")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "external rotation": ("Right External Rotation /\nالدوران الخارجي الأيمن ", "Left External Rotation /\nالدوران الخارجي الأيسر "),
                    "internal rotation": ("Right Internal Rotation /\n الدوران الداخلي الأيمن", "Left Internal Rotation / \nالدوران الداخلي الأيسر"),
                    "flexion":           ("Right shoulder flexion /\n ثني الكتف الأيمن",        "Left shoulder flexion /\n ثني الكتف الأيسر"),
                    "abduction":         ("Right shoulder abductor /\nعضلة فتح الكتف الأيمن",  "Left shoulder abductor /\nعضلة فتح الكتف الأيسر "),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "elbow" and m in ["extension", "flexion"]:
        if (m, r) in elbow_o_movements:
            idx = elbow_o_movements.index((m, r))
            slots = [("O21","P21","O22"),("O23","P23","O24")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "extension": ("Right Triceps /\n عضلات التراي سيبس اليمنى", "Left Triceps /\n عضلات التراي سيبس اليسرى"),
                    "flexion":   ("Right Biceps /\nعضلة الباي سيبس اليمنى",      "Left Biceps /\n عضلة الباي سيبس اليسرى"),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "knee" and m in ["extension", "flexion"]:
        if (m, r) in knee_aa_movements:
            idx = knee_aa_movements.index((m, r))
            slots = [("AA21","AB21","AA22"),("AA23","AB23","AA24")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "extension": ("Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى", "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"),
                    "flexion":   ("Right Hamstring /\nعضلات الفخذ الخلفية اليمنى",   "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    if r == "hip" and m in ["abduction", "adduction"]:
        if (m, r) in hip_ag_movements:
            idx = hip_ag_movements.index((m, r))
            slots = [("AG21","AH21","AG22"),("AG23","AH23","AG24")]
            if idx < len(slots):
                lc, pc, sc = slots[idx]
                txt = {
                    "abduction": ("Right Abductors /\nعضلات الفخذ الخارجية اليمنى", "Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"),
                    "adduction": ("Right Adductors /\nعضلات الفخذ الداخلي اليمنى",  "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"),
                }[m]
                return lc, pc, sc, txt[0], txt[1]
    return None, None, None, None, None


def _build_cells_for_patient(rows, src_ws, test_type, test_types):
    """
    Build the dict of Excel cell assignments for one patient/test_type.
    Returns (cells_dict, movements_list, date_str).
    movements_list is list of (movement, region) tuples that were stored.
    """
    cells = {}
    stored_movements = []

    movements_present = {}
    for row in rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region   = nz_str(src_ws[f"H{row}"].value).lower().strip()
        asym_raw = src_ws[f"S{row}"].value
        if not movement or not region:
            continue
        if test_type == "lower" and region == "trunk":
            continue
        if asym_raw in (None, ""):
            continue
        pct_value, _ = parse_asymmetry(asym_raw)
        if pct_value is None:
            continue
        if len(test_types) > 1:
            row_tt = get_movement_test_type(movement, region)
            if row_tt and row_tt != test_type:
                continue
        key = (movement, region)
        if key not in movements_present:
            movements_present[key] = []
        movements_present[key].append(row)

    # Build ordered movement lists for cell assignment
    if test_type == "lower":
        knee_movements = [k for k in [("extension","knee"),("flexion","knee")] if k in movements_present]
        hip_abd_add    = [k for k in [("adduction","hip"),("abduction","hip")] if k in movements_present]
        hip_flex_ext   = [k for k in [("flexion","hip"),("extension","hip")]   if k in movements_present]
    elif test_type == "upper":
        shoulder_c = [k for k in [("external rotation","shoulder"),("internal rotation","shoulder"),("flexion","shoulder"),("abduction","shoulder")] if k in movements_present]
        shoulder_o = [k for k in [("push","shoulder"),("pull","shoulder")] if k in movements_present]
        elbow_movs = [k for k in [("extension","elbow"),("flexion","elbow")] if k in movements_present]
        hand_movs  = [k for k in [("grip squeeze","hand")] if k in movements_present]
    elif test_type == "full":
        shoulder_c  = [k for k in [("external rotation","shoulder"),("internal rotation","shoulder"),("flexion","shoulder"),("abduction","shoulder")] if k in movements_present]
        elbow_o     = [k for k in [("extension","elbow"),("flexion","elbow")] if k in movements_present]
        knee_aa     = [k for k in [("extension","knee"),("flexion","knee")] if k in movements_present]
        hip_ag      = [k for k in [("abduction","hip"),("adduction","hip")] if k in movements_present]

    # Trunk for lower body
    if test_type == "lower":
        trunk_pct, trunk_weak = calculate_trunk_asymmetry(rows, src_ws)
        if trunk_pct is not None:
            cells["AA21"] = "Trunk Lateral Flexion /\nالثني الجانبي للجذع"
            cells["AB21"] = trunk_pct / 100
            cells["AA22"] = ("Right Sides /\nالجانب الأيمن" if trunk_weak == "Right"
                             else "Left Sides /\nالجانب الأيسر")
            stored_movements.append(("lateral flexion", "trunk"))

    # Process each movement row
    cell_rows = {}
    for row in rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region   = nz_str(src_ws[f"H{row}"].value).lower().strip()
        asym_raw = src_ws[f"S{row}"].value
        if not movement or not region or region == "trunk":
            continue
        if asym_raw in (None, ""):
            continue
        if len(test_types) > 1:
            row_tt = get_movement_test_type(movement, region)
            if row_tt and row_tt != test_type:
                continue
        pct_value, side_char = parse_asymmetry(asym_raw)
        if pct_value is None:
            continue

        if pct_value == 0:
            pct_value = 0.1
            weak_side = None
        else:
            if side_char is None:
                continue
            side_char = side_char.upper()
            weak_side = "Right" if side_char == "L" else ("Left" if side_char == "R" else None)
            if weak_side is None:
                continue

        if test_type == "lower":
            label_cell, pct_cell, side_cell, right_txt, left_txt = get_lower_body_cells(
                movement, region, knee_movements, hip_abd_add, hip_flex_ext)
        elif test_type == "upper":
            label_cell, pct_cell, side_cell, right_txt, left_txt = get_upper_body_cells(
                movement, region, shoulder_c, shoulder_o, elbow_movs, hand_movs)
        elif test_type == "full":
            label_cell, pct_cell, side_cell, right_txt, left_txt = get_full_body_cells(
                movement, region, shoulder_c, elbow_o, knee_aa, hip_ag)
        else:
            continue

        if pct_cell is None:
            continue

        # Keep earliest row (lowest row index = latest measurement in descending export)
        existing_row = cell_rows.get(pct_cell)
        if existing_row is not None and row >= existing_row:
            continue

        cell_rows[pct_cell] = row
        movement_label = get_movement_label(movement, region)
        if label_cell and movement_label:
            cells[label_cell] = movement_label
        cells[pct_cell] = pct_value / 100
        if weak_side is not None:
            cells[side_cell] = right_txt if weak_side == "Right" else left_txt
        if (movement, region) not in stored_movements:
            stored_movements.append((movement, region))

    # Get date from the first row
    date_str = None
    for row in sorted(rows):
        date_val = src_ws[f"C{row}"].value
        if date_val:
            date_str = normalize_test_date(date_val)
            break

    return cells, stored_movements, date_str


# ── main check function ───────────────────────────────────────────────────────

def process_check_file(file_bytes: bytes, gym: str, existing_programs: list[dict], ignored_programs: list[dict] | None = None) -> list[dict]:
    """
    Process an uploaded check Excel file (bytes) and compare against existing
    approved programs from the database.

    existing_programs: list of dicts with keys:
        client_name, test_type, test_date (YYYY-MM-DD), movements

    ignored_programs: list of dicts with keys:
        client_name, test_type, test_date (YYYY-MM-DD), movements
        Tests matching an ignored entry are suppressed unless movement_count increased.

    Returns list of dicts:
        status, patient, external_id, test_type, date, movement_count, old_count
    """
    import io
    src_ws, _wb = _load_worksheet(file_bytes)

    # Build lookup from existing approved programs
    # key: (normalized_client_name, test_type, date_str) → movement_count
    existing_lookup = {}
    # key: (normalized_client_name, test_type, date_str) → {branch, trainer_name, dispatch_date}
    existing_details_lookup: dict[tuple, dict] = {}
    # key: (normalized_client_name, test_type) → asymmetry_values from latest approved
    prev_asymmetries_lookup: dict[tuple, dict] = {}
    for p in existing_programs:
        norm_name = re.sub(r"\s+", " ", str(p["client_name"]).strip())
        tt = p["test_type"]
        date_str_p = normalize_test_date(p["test_date"])
        key = (norm_name, tt, date_str_p)
        mv = p.get("movements", 0)
        existing_lookup[key] = len(mv) if isinstance(mv, list) else (mv or 0)
        existing_details_lookup[key] = {
            "branch": p.get("branch") or None,
            "trainer_name": p.get("trainer_name") or None,
            "dispatch_date": p.get("dispatch_date") or None,
        }
        # Track the latest date per (name, test_type) for comparison
        pa_key = (norm_name, tt)
        av = p.get("asymmetry_values")
        if av:
            existing_date = prev_asymmetries_lookup.get(pa_key, {}).get("_date", "")
            if date_str_p and date_str_p > existing_date:
                prev_asymmetries_lookup[pa_key] = {**av, "_date": date_str_p}

    # Build lookup for ignored tests: key → movement_count at time of ignore
    # A test is suppressed if it matches the key AND movement_count hasn't increased
    ignored_lookup: dict[tuple, int] = {}
    for p in (ignored_programs or []):
        norm_name = re.sub(r"\s+", " ", str(p["client_name"]).strip())
        tt = p["test_type"]
        date_str_p = normalize_test_date(p["test_date"])
        key = (norm_name, tt, date_str_p)
        mv = p.get("movements", 0)
        ignored_lookup[key] = len(mv) if isinstance(mv, list) else (mv or 0)

    # Collect rows per patient per date (prevents cross-date full-body detection)
    # patients_rows: { patient_name: { date_str: [row, ...] } }
    patients_rows = {}
    patient_external_ids = {}
    for row in range(2, src_ws.max_row + 1):
        name_val = re.sub(r"\s+", " ", nz_str(src_ws[f"A{row}"].value).strip())
        if not name_val:
            continue
        date_val = src_ws[f"C{row}"].value
        date_key = normalize_test_date(date_val) or ""
        if name_val not in patients_rows:
            patients_rows[name_val] = {}
            ext_id = nz_str(src_ws[f"B{row}"].value).strip()
            patient_external_ids[name_val] = ext_id if ext_id else "N/A"
        if date_key not in patients_rows[name_val]:
            patients_rows[name_val][date_key] = []
        patients_rows[name_val][date_key].append(row)

    new_tests = []

    for patient_name, dates_rows in patients_rows.items():
        for _date_key, rows in dates_rows.items():
            test_types = detect_test_type(rows, src_ws)
            if isinstance(test_types, str):
                test_types = [test_types]

            for test_type in test_types:
                movements_present = {}
                movements_stored = {}
                rows_with_stored_movements = set()

                for row in rows:
                    movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
                    region = nz_str(src_ws[f"H{row}"].value).lower().strip()
                    asym_raw = src_ws[f"S{row}"].value

                    if not movement or not region:
                        continue

                    # Trunk is calculated separately for lower body
                    if test_type == "lower" and region == "trunk":
                        trunk_pct, _ = calculate_trunk_asymmetry(rows, src_ws)
                        if trunk_pct is not None:
                            if ("lateral flexion", "trunk") not in movements_stored:
                                movements_stored[("lateral flexion", "trunk")] = trunk_pct
                        continue

                    if asym_raw in (None, ""):
                        continue
                    pct_value, _ = parse_asymmetry(asym_raw)
                    if pct_value is None:
                        continue

                    if len(test_types) > 1:
                        row_test_type = get_movement_test_type(movement, region)
                        if row_test_type and row_test_type != test_type:
                            continue

                    would_be_stored = False
                    if test_type == "lower":
                        if region == "knee" and movement in ("extension", "flexion"):
                            would_be_stored = True
                        elif region == "hip" and movement in ("abduction", "adduction", "flexion", "extension"):
                            would_be_stored = True
                    elif test_type == "upper":
                        if region == "shoulder" and movement in ("external rotation", "internal rotation", "flexion", "abduction", "push", "pull"):
                            would_be_stored = True
                        elif region == "elbow" and movement in ("extension", "flexion"):
                            would_be_stored = True
                        elif region == "hand" and movement == "grip squeeze":
                            would_be_stored = True
                    elif test_type == "full":
                        if region == "shoulder" and movement in ("external rotation", "internal rotation", "flexion", "abduction"):
                            would_be_stored = True
                        elif region == "elbow" and movement in ("extension", "flexion"):
                            would_be_stored = True
                        elif region == "knee" and movement in ("extension", "flexion"):
                            would_be_stored = True
                        elif region == "hip" and movement in ("abduction", "adduction"):
                            would_be_stored = True

                    if not would_be_stored:
                        continue

                    key = (movement, region)
                    if key not in movements_present:
                        movements_present[key] = []
                    movements_present[key].append((abs(pct_value), row))

                for key, value_row_pairs in movements_present.items():
                    max_value, max_row = max(value_row_pairs, key=lambda x: x[0])
                    movements_stored[key] = max_value
                    rows_with_stored_movements.add(max_row)

                if len(movements_stored) == 0:
                    continue

                date_val = None
                for row in sorted(rows_with_stored_movements):
                    date_val = src_ws[f"C{row}"].value
                    if date_val:
                        break

                if not date_val:
                    continue

                date_str = normalize_test_date(date_val)
                movement_count = len(movements_stored)

                lookup_key = (patient_name, test_type, date_str)
                is_new = lookup_key not in existing_lookup
                old_count = existing_lookup.get(lookup_key, 0)

                if is_new or movement_count > old_count:
                    # Suppress if this exact test was ignored and hasn't gained new movements
                    ignored_count = ignored_lookup.get(lookup_key)
                    if ignored_count is not None and movement_count <= ignored_count:
                        continue
                    # Build template cell data for program generation
                    cells_dict, stored_movs, _ = _build_cells_for_patient(rows, src_ws, test_type, test_types)
                    # Build asymmetry_values dict (movement_key -> pct 0-100) for future comparisons
                    from program_builder import _CELL_GROUPS, _movement_key
                    av: dict[str, float] = {}
                    for lbl_col, pct_col in _CELL_GROUPS:
                        for row_num in (21, 23, 25, 27):
                            lv = str(cells_dict.get(f"{lbl_col}{row_num}", "")).strip()
                            pv = cells_dict.get(f"{pct_col}{row_num}")
                            if lv and pv is not None:
                                mk = _movement_key(lv)
                                if mk:
                                    av[mk] = round(abs(float(pv)) * 100, 4)
                    cells_data = {
                        "cells": cells_dict,
                        "movements": [list(m) for m in stored_movs],
                    }
                    # Get previous asymmetries for this patient+test_type
                    pa_key = (patient_name, test_type)
                    prev_av_raw = prev_asymmetries_lookup.get(pa_key, {})
                    prev_av = {k: v for k, v in prev_av_raw.items() if k != "_date"}
                    # If the previous approved test is the same date as the current test,
                    # it's the same test with more movements added — don't compare (no red font)
                    prev_date = prev_av_raw.get("_date", "")
                    if prev_date == date_str:
                        prev_av = {}
                    # For updated tests, carry over the existing branch/trainer/dispatch_date
                    existing_details = existing_details_lookup.get(lookup_key, {})
                    new_tests.append({
                        "status": "NEW" if is_new else "UPDATED",
                        "patient": patient_name,
                        "external_id": patient_external_ids.get(patient_name, "N/A"),
                        "test_type": test_type,
                        "date": date_str,
                        "movement_count": movement_count,
                        "old_count": old_count,
                        "cells_data": cells_data,
                        "asymmetry_values": av,
                        "prev_asymmetries": prev_av if prev_av else None,
                        "existing_branch": existing_details.get("branch"),
                        "existing_trainer_name": existing_details.get("trainer_name"),
                        "existing_dispatch_date": existing_details.get("dispatch_date"),
                    })

    if _wb:
        _wb.close()
    return new_tests
