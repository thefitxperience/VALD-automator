"""
Generates the monthly / weekly Excel report from the template files.
Fills each branch sheet starting at row 7 with approved programs.
"""
import io
import os
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Template files are copied into the same directory as this file in the container
TEMPLATE_MAP = {
    "Body Masters": os.path.join(BASE_DIR, "Month YEAR - Body Masters.xlsx"),
    "Body Motions": os.path.join(BASE_DIR, "Month YEAR - Body Motions.xlsx"),
}

BRANCH_ORDER = {
    "Body Masters": [
        "RUH - Al Malaz", "RUH - Al Massif", "RUH - Al Aarid", "RUH - Al Sahafa",
        "RUH - Al Wadi", "RUH - Eshbilia", "RUH - Muzahmiyah", "RUH - Rabwa",
        "RUH - Salam", "RUH - Swaidi", "RUH - Takhasousi", "RUH - Al Badia",
        "RUH - Al Fayha", "RUH - Al Khaleej", "RUH - Al Kharj", "RUH - Al Nahda",
        "RUH - Badr", "RUH - Ezdehar", "RUH - Murooj", "RUH - Shubra",
        "DMM - Al Athir", "DMM - Al Jameyeen", "DMM - Hufof", "DMM - Khobar",
        "JED - Hamadania", "JED - Al Rawdah", "JED - Makkah",
        "JED - Obhor - Al Amwaj", "JED - Obhor - Al Sheraa",
        "ALQ - Al Rass", "ALQ - Buraidah", "ALQ - Unaizah",
        "MED - Shouran", "MED - Taiba",
        "Uhud", "AlUla", "Al Mubaraz", "Hafr El Batin", "Tabuk", "Najran",
        "Khamis Mushait", "Hail",
    ],
    "Body Motions": [
        "RUH - Al Malaz", "RUH - Al Sahafa", "RUH - Al Aarid", "RUH - Al Fayha",
        "RUH - Al Uraija", "RUH - Badr", "RUH - Al Badia",
        "JED - Al Basateen", "JED - Al Faisaliyah", "JED - Al Naeem",
        "DMM - Al Faisaliyah", "DMM - Al Jalawiah", "DMM - Al Nada",
        "ALQ - Buraidah", "ALQ - Unaizah",
        "Al Ahsaa", "AlUla", "Tabuk",
    ],
}

TRAINER_ORDER = {
    "Body Masters": [
        # RUH - Al Malaz
        "Omar Ismael","Mohaned Ahmad","Yassine Abbich","Chady Taha Abdel Rahman",
        "Hicham Mahboub","Omar Saoud","Adnan Abdallah Aljahny","Adnan Abdallah Al Jahni",
        "Abdallah Kamal Ahmad Sayid","Mohamed Elshabrawy","Mahmoud Bakry","Trainer Ismail",
        # RUH - Al Massif
        "Mohamad Touri","Jayson Atlano","Mohammad Souud","Oussama Elbaz","Sleiman Al Tayyar",
        "Ossama Almaslout","Alaa Fathy","Youssef Lekal","Anouar Aghmour","Amer Abid",
        "Mahmoud Abdel Rahim Senussi Kame","Mahmoud Abdel Rahim Senussi",
        "Hany Mohamed Yousif Elhadad","Ahmed Keshk","Mohammed Hida","Ahmed Shoeeb",
        "Saleh Al Abed","Asmah Choukeni",
        # RUH - Al Aarid
        "Diyaa Al Jraydan","Mhamad El Ebri","Chehab Al Dine Al Salihi",
        "Captain Housein Ali Joumaa","Mahmoud Abed El Rahim","Khalid Mouazen",
        "Mourad Belgacem","Moustafa Nasser Abd Elkader","Amr El Sayed Abdelhamid",
        "Mahmoud Yosri Mahmoud Mahfouz",
        # RUH - Al Sahafa
        "Hamza Sanbar","Fawzi Baltaji","Magomed Ansarov","Ahmad Fadel","Hisham M. Badih",
        "Mohamad Zakariah","Ayman N. Al Chahawi","Diya Y. Aljirdan","Ahmad H. Abed Al Latif",
        "Rachid Sakouti","Mohamad Fouad G. Amer","Rrnniel Villaverde","Zakaria Yahia",
        "Akram M. Hussin","Zakaria Jawadi","Youssef Noreddine","Renelle Vilaverdi",
        # RUH - Al Wadi
        "Shamsudheen Paramban","Abdellatif Zamzim","Ahmed Hamid Abdellatif Nogalla",
        "Ibrahim Mohamed Fadel Abduljalil Barakat","Bander Albalawi","El Kaddioui Adil",
        "Omar Mayhoub","Abdou Wahed Zraidi",
        # RUH - Eshbilia
        "Houssam Ali Elsheshtawy","Bilel Azhar Yakoubi","Saud Turki Alqasir",
        "Mohamd Refai Mohamd","Mohamed Ibrahim Ibrahim Abouelrous","Ibrahim Mostafa",
        "Marwan Al-Jebali",
        # RUH - Muzahmiyah
        "Khalifa El Nahrawy","Abdssamad Aballa","Nabil Essahraoui",
        # RUH - Rabwa
        "Aref Hebatala","Alaa Eddine Rtili","Abderrazak Sihabi","Mostafa Ashour Elkhouly",
        # RUH - Salam
        "Hicham Mouhime","Abdellatif Zamzam","Fady Elsayed","Mohamed Kamal",
        # RUH - Swaidi
        "Nasser Mansour Nasser Alessa","Mohammed Nshbat","Naoufel Jabnouni",
        "Khalid Chatir","Mohammed Korchi",
        # RUH - Takhasousi
        "Saeed Al Adrham","Mahmoud Nagib","Taib Jadia","Khalid Nokracchi",
        "Wasem Seed Alhajree","Salman Altayyar","Ahmed Abdelrahman","Fikri Acharif",
        "Abdenabi Ktraouach","Ahmed Adel Abdellatif Farag",
        # RUH - Al Badia
        "Adil Chahmi","El Mehdi Adnani","Khalifa Hassan Khalifa Nahrawy",
        "Mohamad Mahdim Alqahtani","Amine Korchi","Abderrazak Sihabi","Mahmoud Fotoh",
        "Adil Elqady","Mohamed Saeed","Selim Ben Abdelkader",
        # RUH - Al Fayha
        "Faraj Lakaira","Ibrahim Pokoti","Hicham Mouhime","Mahmoud Al Mahdi",
        "Ahmad Hamdi Knawi","Kamal Al-Rayes","Asser Al-Sheikh","Hamza Aissaoui",
        # RUH - Al Khaleej
        "Chouaib Suilmi","Ahmad Ramadan","Karim Hosny Ahmad","Mohamed Moustafa Ahmad",
        "Oussama Ait Hammou Obrahim","Zakaria Elouali","Elsayed Fouad",
        "Mohamed Hassan Ali Hassan","Sameh Houssein Ali Abed Al-Wahid",
        # RUH - Al Kharj
        "Mohamed Hamdan Mohamed","Nabil Essahraoui","Mahmoud Mohamed",
        "Yassine El Barnoussi","Adel Ali Mohamed","Boujemaa Elakkad",
        # RUH - Al Nahda
        "Youssef Al Habib Trabelsi","Mhamed Oulahssine","Rhey Anthony T. Munez",
        "Majid Abdessamad","Ramy Alla Tawfeak Ahmed","Achraf Fadlallah","Soufian Bija",
        "Raqan Abdel Mehsen Al-Qahtani","Hicham Toualbi","Abed El Hakim Sahou",
        # RUH - Badr
        "Mohammed Mubarak Mohammed Aldawsari","Mahmoud Farouk Abdel Karim Badawi",
        "Khalled Saeed Saeed Alhasany","Mohammed Bstan Abd Aljwad",
        "Kamel Mohammed Kamel","Adawi Abdelmounnmadawi",
        # RUH - Ezdehar
        "Wajih Kamal Matoussi","Tarik Maarouf","Mohsen Hadikhamis",
        "Ahmed Maher Ahmed Nasrelden","Hesham Rady Ismail Rady","Ibrahim Ali Ismaiel Ibrahim",
        # RUH - Murooj
        "Abdelaziz El Maydy","Amr Ahmed Abdellatif Rabia","Cherif Bin Ahmed Khlifi",
        # RUH - Shubra
        "Ahmed Maher Elsayed Abdelaal","Mohammed Abdullah Alquwaya","Hicham Oulias",
        "Rachid Choukhmane","Mohammed Mahdi Adham Nabhan","Osman Nagi Osman",
        "Abdessamade Moudakkir","Tariq Abdulrahman Basharahil",
        # DMM - Al Athir
        "Abdullah Abdulrahman","Abdulrahman Ali","Ibrahim Salama","Rashid Elbouazizi",
        "Mohamed Elsayed Elzeftawy","Alaa El Din Mekhtar","Ammar Ali","Jamal Halabi Assiri",
        # DMM - Al Jameyeen
        "Nader Maayouf Saleh","Hesham Adel Almarshoud","Mohamed Khalid Manni",
        "Mohamed Abdelsalam Deyab","Hamza El Moumine","Mohamed Araby","Shahid Elkhalfy",
        "Mohammed Abdelhamid","Ashraf Ezzeldine","Ibrahim Babaker","Mohamad Ahmad",
        "Ali Ahmed Alkhadhir",
        # DMM - Hufof
        "Khalid Karmou","Ahmad Abdelgafour Omar Ahmad","Abdelrahman Mohamed",
        "Abdul Raouf Mustafa",
        # DMM - Khobar
        "Radwan Mohamed Radwan","Abbas Okasha","Khalid Mahmoud Saleh","Abed AlWasi",
        "Mahmoud Hassan","Shehab Ashraf Mohamady Ibrahim","Ayoub Chatat","Ahmad Al Dokhi",
        "Joesel Bodota Casimiro","Mohamad Maani","Rabana Gabif",
        # JED - Hamadania
        "Abdullah Mohammed Alhamdni","Musa Eltom Ahmed","Mohamed Ibrahim Mohamed",
        "Abd Elramhan Mokhtar","Hassan Bargougui",
        # JED - Al Rawdah
        "Ahmed Elsayed Algammal","Muhmoud Ayman Saad","Youssef Souayah",
        "Mohamed Fathi Abdulsattar","Abdallah Kandaly","Sofiane Sadiq","Ibrahim Nasr",
        "Sami Serrar","Khaled Al-Masouri","Tawfik Jukh","Moez Mustafa","Soliman Bakar",
        "Oussama Amzil",
        # JED - Makkah
        "Mohamed Gamea","Omar Al Shankety","Mohamed Ait Ayad","Mahmoud Al Bihery",
        "Adel Mohammed Abdullah Al-Daos","Khaled Al Barkawi","Imad Al-Zahrani",
        # JED - Obhor - Al Amwaj
        "Hesham Galal Mahmoud","Karim Abdala Wassal","Mohammed Joharji","Mohammed Mustafa Himsi",
        # JED - Obhor - Al Sheraa
        "Abdelrahman Faris","Ibrahim Saleh","Nawaf Alanazi","Mostafa Hamada",
        "Tarak Othman","Abdelrahman Kamal","Bader Eddine","Ahmed Al Gammal",
        "Ayoub Benzahra","Abed Al Menhem Mhamad","Mohamad Rida Irabi",
        "Badereddine Bouzahar","Abed El Rahman Taha",
        # ALQ - Al Rass
        "Yousef Bou Kentar","Hicham Hemsi",
        # ALQ - Buraidah
        "Kamis Zamzam","Yasser Mahmoud","Mouad Omir","Oussema Boussaid","Orley Garcia",
        # ALQ - Unaizah
        "Karim Maarori","Mohamed Othman Gazy","Mohamed Jaouat","Mohamed Amzil",
        "Ahmed Tarek Kholafa",
        # MED - Shouran
        "Sheref Ashraf Mohamed","Mohamed Soliman Ragab","Abdel Ghafour Bahja",
        "Lakhdari Abdellatif","Anas Mohmmed Alshanqiti","Ahmed Mohamed Helmy Tohamy",
        "Adham Ahmed Abdelsttar","Ahmed Abdallah Elsayed Ibrahim",
        "Hassan Adel Almughamisi","Amjad Hussein Abbs Abdulaal",
        # MED - Taiba
        "Abdelaziz Mohi Eldin Elkhouli","Mostafa Mahmoud Sayed Mohamed",
        "Amgad Hussein Abbas Abdulaal","Abdulwahab Ahmed Saad Al Ghamdi",
        "Rabah Salamah Alhujuri",
        # Uhud
        "Abdallah Mohamed Ali","Hamza Berahhou","Wael Mohammed","Moulay Al Hussein",
        "Hassan Ait Hmida","Ahmad Al-Dawkhi","Hussein Elsayed","Anas Al-Makenzi",
        "Mohamed Rabhi","Enrique Jr Ravallo-Macorol","Mohamed Abdellatif Radwan",
        "Hamza Said","Hicham Hassi","Mohamed Abdel Moeti","Houssem Saggedi",
        # AlUla
        "Yassin Mbarek","Ahmed Khaled","Ahmed Khaled El Said","Moamed Ata","Mohannad Hani",
        "Ahmed Elsayed","Mohamad Amine Azhar","Ahmed Abdelmaqsoud","ElHassan Nagib",
        # Al Mubaraz
        "Sabry Ali Abdelmonem Ezzah","Abdallah Lazam Faisal Al-Asadi",
        "Ammar Tarek Fathi Ibrahim Ali Mahmoud","Mohamed Fouad Amer","Hamza Moustaghfir",
        "Radi Gamal Osman Hussein","Samer Mohammed Abas Kamber","Khaled Ezzat",
        "Abd Elrahem Al Ouina","Amar Fathi",
        # Hafr El Batin
        "Mohamed Ibrahim Abdelaal","Ahmed Saad Elmamlouk","Mohamed Abdel Manaam Yakout",
        "Mohamed Achraf Abed Al-Mawjoud Mohamed","Mahmoud El-Sayed Abdelwareth Radwan",
        "Mohamad Ashraf","Abdel Rahman Hamad","Ahmad Al Hussein",
        # Tabuk
        "Mohammad Nour Al-Din Awad Mohammad","Miloud Baali","Mohammad Ismail",
        "Tamer Sharkas","Mostafa Hamada","Anas Jouhari",
        # Najran
        "Said Ennasr","Ahmed Mohamed Ahmed Fadol","Wadhah Ahmed Ahmed Mahyub",
        "Ibrahim Refat Abdellatif","Abdaleleh Naser Ahmed",
        # Khamis Mushait
        "Mohammed Ayed","Mustapha Selaoui","Ibrahem Zahran","Ali Alkobebi","Hamdy Helal",
        "Eslam Rezk","Amin Haron",
        # Hail
        "Ahmad Chiboub","Ali Mohamed Ali Mohamed","Bassam Ahmad Al-Salhani",
        "Abdallah Laghzal","Hatem Kamal Hatiba","Wagih Al-Saadani",
    ],
    "Body Motions": [
        # RUH - Al Malaz
        "Abir Salloum","Faemeh Al Seraji","Ikram Gueddes","Latoya Fantisi",
        "Mariam Jahin","Moroj Samir Abdel Hamid","Nadine Ghaleb","Sarah Shaabane",
        "Loel Du Bruyn","Taghrid Khaled","Shayma Hanashi","Nwabisa Liwane",
        "Areej Hajj Abdallah","Hadir Nasr Mohamed","Taif Saad Alahmari","Zouhaira Al-Omrani",
        # RUH - Al Sahafa
        "Shefa Ghazi","Manel Manasria","Abeer Al Saloum","Rania Manita",
        "Kholoud Khaled","Guesmi Kholoud","Nashwa Attia Elbendary","Sirine Mohamed Hamdi",
        "Maram Tlili","Sara Said","Malek Belguith","Safaa Karim",
        # RUH - Al Aarid
        "Amal Al Chiboub","Hidaya Moubarak","Safa Al Saggadi","Warda Al Hajji",
        "Maryam Al-Qawsari","Fawz Jibran Al-Qahtani","Malak Al Zahabi",
        "Nada Mohamed Khalifa","Chaza Ben Miled","Shaden Omar",
        # RUH - Al Fayha
        "Walaa Kamel","Hind Al Ahmad","Asmaa Adel","Maram Al Sousou",
        "Asmaa Ahmed","Ebtihal Al Kathiri","Feryel Ben Nacer",
        "Al Anoud Saleh Al-Qablan","Khawla Saif",
        # RUH - Al Uraija
        "Basnat Abdel Qawi","Sanae Hassel","Khouloud Wanna",
        "Hadiya Al-Mohammad","Aya Al-Jallouli","Rania Bahloul",
        # RUH - Badr
        "Shorouq Osama Metwaly Ibrahim","Arwa Ahmed Baflah","Hadeer Ibrahim Gamal",
        "Rihem Hossam","Sarah Ghoneim","Al-Hanouf Al-Khurayan",
        "Amal Al-Bjeoui","Seham Hussein","Maysaa Mowafaq Al-Aqram",
        # RUH - Al Badia
        "Takwa Abdelly","Hasna AlKayyal","Souha Al-Bajawi","Hanin Barrani",
        "Safa Radif","Maram Farhawi","Rajae Al-Jabouji",
        "Ghofran Bint Kamal Aouni","Nermin Hadi Debaya","Nawres Khordeni",
        # JED - Al Basateen
        "Raghad Al Bir Qaddar","Iman Kouisseh","Salwa Mlaiki","Sabrine Al Amine",
        "Mouna Farhat","Moutidis Ofantisi","Shurooq Amer","Reyouf Saad",
        "Ranim Brahmi","Maha Fayez Al Ahmari","Salma Jadli",
        # JED - Al Faisaliyah
        "Hala Saied","Kholod Khaled","Wiem Dridi","Shourooq Al Jundi",
        "Ines Rezgui","Hajer Jamal Al-Deen Khalifa Ali","Mariam Baomar","Chourouk Sherif",
        # JED - Al Naeem
        "Sara Mahdi","Siwar Al Habib","Shayma Abbes","Nahla Al Hajri",
        "Afraa Almasri","Lamia Al Amer","Diaa Sobhi","Maysem Mohamed",
        "Asayel Sultan","Nadine Shawky","Mayssa Bin Mohamad Nasser Mejri",
        # DMM - Al Faisaliyah
        "Rasha Hamid","Amira Omar","Soulaima Hassan","Cyrine Ali",
        "Sirine Jawadi","Moulouk Al Lawi","Nourhan Hosny","Al Bandari Al Harbi",
        "Mirna Mostafa","Mariem Laamiri","Rahma Qays",
        # DMM - Al Jalawiah
        "Asmaa Mahmoud","Basma Baza","Mennatllah Saad","Asmaa Ashraf",
        "Chaima Gharsalli","Fatima Abdallah","Emna Ben Chalbia",
        "Hebatallah Ali","Aya Maayoufi",
        # DMM - Al Nada
        "Amal Al-Fahat Ghazi Al-Ruwaili","Chaima Boutita","Asmaa Shehata",
        "Sarah Rasheeq","Nawal Akermi",
        # ALQ - Buraidah
        "Rawiya Alarousi Mabrouk","Anfal Nasser Alhusseini","Nourhan Ehab Amarah",
        "Nouran Emad Jad","Lulu Alzaaj","Malath Suleiman Alshamasi",
        "Chayma Hannachi","Mai Mosleh","Ines Souissi","Anwar Ben Rhayem",
        "Raja Reeba","Aya Mahmoud Abed Al-Wanis","Eilaf Mohamed Al-Fayizi",
        # ALQ - Unaizah
        "Rana Mohammad Hassan Aaref","Roua Al-Hashemi Bin Naeimah","Yosra Al-Mzoughi",
        "Esraa Labib Noomani","Asmaa Aalaa Al-Din","Samar Lajnef","Rim Chabi",
        # Al Ahsaa
        "Iman Aboushosha","Marwa Almudyni","Afnan Khalid","Mona Maatouq",
        "Sohaila Refaat","Samar Mohammed","Esraa Idris","Omaymah Youssef",
        "Habiba Alghareeb","Amina Mohammad Mosbeh","Nour Toumi",
        # AlUla
        "Khouloud Suibgui","Rehab Naji Al-Nasri","Hanin Barrani",
        "Bayan Al-Anzi","Amal Saleh","Walaa Ahmed",
        # Tabuk
        "Rasha Abu Hassan","Nada Ahmed Mohamed","Ishraf Mohamed Al-Zaghdoudi",
        "Sahar Mrad Berguiga","Marwa Jendoubi","Bassant Majed Eid",
        "Amal Saad","Hanane Bankroum",
    ],
}

TEST_TYPE_LABELS = {
    "upper": "Upper Body",
    "lower": "Lower Body",
    "full": "Full Body",
}


def _week_range(year: int, month: int, week_number: int):
    """Return (start_date, end_date) for week 1-4/5 of a month."""
    first = date(year, month, 1)
    # week 1 = days 1-7, week 2 = 8-14, week 3 = 15-21, week 4 = 22-end
    start_day = (week_number - 1) * 7 + 1
    end_day = start_day + 6
    # Clamp end_day to last day of month
    if month == 12:
        next_month_first = date(year + 1, 1, 1)
    else:
        next_month_first = date(year, month + 1, 1)
    last_day = (next_month_first - timedelta(days=1)).day
    end_day = min(end_day, last_day)
    return date(year, month, start_day), date(year, month, end_day)


def _copy_row_style(src_ws, src_row: int, dst_ws, dst_row: int, max_col: int):
    """Copy cell styles from a reference row to a new data row."""
    for col in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row, column=col)
        dst_cell = dst_ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format


def generate_report(
    gym: str,
    programs: list[dict],
    period_type: str,   # "monthly" | "weekly"
    year: int,
    month: int,
    week_number: int | None = None,
    start_day: int | None = None,
    end_day: int | None = None,
    report_date: date | None = None,
) -> bytes:
    """
    Build a report Excel file and return its bytes.

    programs: list of dicts from Supabase with keys:
        branch, client_id, client_name, test_type, test_date,
        trainer_name, dispatch_date
    """
    template_path = TEMPLATE_MAP.get(gym)
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Determine date filter
    if period_type == "monthly":
        import calendar as _cal
        last_day = _cal.monthrange(year, month)[1]
        period_start = date(year, month, max(1, start_day or 1))
        period_end = date(year, month, min(last_day, end_day or last_day))
    else:
        if week_number is None:
            raise ValueError("week_number required for weekly report")
        period_start, period_end = _week_range(year, month, week_number)

    # Filter programs by dispatch_date in period
    def in_period(p):
        dd = p.get("dispatch_date")
        if not dd:
            return False
        if isinstance(dd, str):
            try:
                dd = date.fromisoformat(dd)
            except ValueError:
                return False
        return period_start <= dd <= period_end

    filtered = [p for p in programs if in_period(p)]

    # Group period-filtered programs by branch (used for branch sheets)
    by_branch: dict[str, list] = {}
    for p in filtered:
        branch = p.get("branch", "")
        if branch not in by_branch:
            by_branch[branch] = []
        by_branch[branch].append(p)

    # Monthly totals per branch (full month, regardless of period filter)
    # Used for REPORT 2 column C in weekly/custom reports
    import calendar as _cal
    month_last = _cal.monthrange(year, month)[1]
    month_start = date(year, month, 1)
    month_end = date(year, month, month_last)

    def in_full_month(p):
        dd = p.get("dispatch_date")
        if not dd:
            return False
        if isinstance(dd, str):
            try:
                dd = date.fromisoformat(dd)
            except ValueError:
                return False
        return month_start <= dd <= month_end

    monthly_by_branch: dict[str, int] = {}
    for p in programs:
        if in_full_month(p):
            branch = p.get("branch", "")
            monthly_by_branch[branch] = monthly_by_branch.get(branch, 0) + 1

    # Monthly totals per trainer (full month) for REPORT sheet
    monthly_by_trainer: dict[str, int] = {}
    for p in programs:
        if in_full_month(p):
            trainer = p.get("trainer_name", "") or ""
            monthly_by_trainer[trainer] = monthly_by_trainer.get(trainer, 0) + 1

    # Load template
    wb = load_workbook(template_path, data_only=False)

    rpt_date = report_date or date.today()
    # For the summary sheets, use period_end when a custom date range is set
    summary_date = period_end if (start_day or end_day) else rpt_date

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip summary sheets
        if sheet_name in ("REPORT", "REPORT 2"):
            # Update report date cell (B3)
            ws["B3"] = summary_date
            # For weekly/custom reports, fill column C from C7 with branch counts
            # (monthly uses a formula that counts from branch sheets directly)
            is_partial = period_type == "weekly" or bool(start_day or end_day)
            if is_partial and sheet_name == "REPORT 2":
                branch_order = BRANCH_ORDER.get(gym, [])
                for idx, branch in enumerate(branch_order):
                    count = monthly_by_branch.get(branch, 0)
                    cell = ws.cell(row=7 + idx, column=3)
                    cell.value = None
                    cell.value = count
            if is_partial and sheet_name == "REPORT":
                trainer_order = TRAINER_ORDER.get(gym, [])
                for idx, trainer in enumerate(trainer_order):
                    count = monthly_by_trainer.get(trainer, 0)
                    cell = ws.cell(row=7 + idx, column=3)
                    cell.value = None
                    cell.value = count
            continue

        # Branch sheet — update report date
        ws["B3"] = rpt_date

        branch_programs = by_branch.get(sheet_name, [])
        if not branch_programs:
            continue

        # Find where to start writing (row 7 and down)
        start_row = 7
        max_col = ws.max_column

        # Use row 7 as style template if it has data (from previous use), else row 5 headers
        style_ref_row = 7

        for i, prog in enumerate(branch_programs):
            dest_row = start_row + i
            _copy_row_style(ws, style_ref_row, ws, dest_row, max_col)

            client_id = prog.get("client_id") or ""
            client_name = prog.get("client_name", "")
            test_type_label = TEST_TYPE_LABELS.get(prog.get("test_type", ""), "")
            trainer = prog.get("trainer_name") or ""
            test_date = prog.get("test_date")
            dispatch_date = prog.get("dispatch_date")

            # Format dates
            if isinstance(test_date, str):
                try:
                    test_date = date.fromisoformat(test_date)
                except ValueError:
                    pass
            if isinstance(dispatch_date, str):
                try:
                    dispatch_date = date.fromisoformat(dispatch_date)
                except ValueError:
                    pass

            full_name = f"{client_name} - {test_type_label}" if test_type_label else client_name

            ws.cell(row=dest_row, column=1, value=client_id)
            ws.cell(row=dest_row, column=2, value=full_name)
            ws.cell(row=dest_row, column=3, value=trainer)
            ws.cell(row=dest_row, column=4, value=test_date)
            ws.cell(row=dest_row, column=5, value=dispatch_date)

            # Format date cells
            ws.cell(row=dest_row, column=4).number_format = "DD/MM/YYYY"
            ws.cell(row=dest_row, column=5).number_format = "DD/MM/YYYY"

            # Column F: mark late uploads (test done in a previous month)
            if isinstance(test_date, date) and test_date.month != month:
                ws.cell(row=dest_row, column=6, value="Late Upload")

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
