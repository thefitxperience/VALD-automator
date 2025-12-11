import sys
import os
import shutil
import re
from openpyxl import load_workbook
import subprocess
import xlwings as xw
from datetime import datetime
import json


def get_movement_test_type(movement, region):
    """
    Determine which test type a specific movement belongs to.
    Returns: 'upper', 'lower', or 'full'
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    # Upper body movements
    upper_movements = [
        (r == "shoulder"),
        (r == "hand"),
        (r == "elbow" and m in ["extension", "flexion"])
    ]
    
    # Lower body movements
    lower_movements = [
        (r == "trunk"),
        (r == "knee"),
        (r == "hip" and m in ["flexion", "extension", "abduction", "adduction"])
    ]
    
    if any(upper_movements):
        return 'upper'
    elif any(lower_movements):
        return 'lower'
    
    return None


def detect_test_type(patient_rows, src_ws):
    """
    Analyze the actual movements and regions in the data to determine test type.
    Returns: 'upper', 'lower', 'full', or list like ['upper', 'lower']
    """
    movements_regions = set()
    
    for row in patient_rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region = nz_str(src_ws[f"H{row}"].value).lower().strip()
        if movement and region:
            movements_regions.add((movement, region))
    
    # Categorize movements
    upper_indicators = {
        ('external rotation', 'shoulder'),
        ('internal rotation', 'shoulder'),
        ('flexion', 'shoulder'),
        ('abduction', 'shoulder'),
        ('push', 'shoulder'),
        ('pull', 'shoulder'),
        ('grip squeeze', 'hand')
    }
    
    elbow_indicators = {
        ('extension', 'elbow'),
        ('flexion', 'elbow')
    }
    
    lower_indicators = {
        ('lateral flexion', 'trunk'),
        ('flexion', 'hip'),
        ('extension', 'hip')
    }
    
    knee_indicators = {
        ('extension', 'knee'),
        ('flexion', 'knee')
    }
    
    hip_abd_add_indicators = {
        ('abduction', 'hip'),
        ('adduction', 'hip')
    }
    
    # Check what's present
    has_upper = any(mr in movements_regions for mr in upper_indicators)
    has_elbow = any(mr in movements_regions for mr in elbow_indicators)
    has_knee = any(mr in movements_regions for mr in knee_indicators)
    has_hip_abd_add = any(mr in movements_regions for mr in hip_abd_add_indicators)
    has_lower_specific = any(mr in movements_regions for mr in lower_indicators)
    
    # Detection logic
    test_types = []
    
    # Full Body: Has upper (shoulder) + elbow + knee + hip abduction/adduction (but NOT hip flexion/extension or trunk)
    if has_upper and has_elbow and has_knee and has_hip_abd_add and not has_lower_specific:
        return 'full'
    
    # Check if we have upper movements
    has_any_upper = has_upper or has_elbow
    
    # Check if we have lower movements
    has_any_lower = has_knee or has_hip_abd_add or has_lower_specific
    
    # If we have both upper and lower, return both as separate tests
    if has_any_upper and has_any_lower:
        return ['upper', 'lower']
    
    # If only upper
    if has_any_upper:
        return 'upper'
    
    # If only lower
    if has_any_lower:
        return 'lower'
    
    # Default
    return 'upper'


def get_template_for_test_type(test_type, gym_folder, base_dir):
    """
    Get the template path for a specific test type.
    """
    body_type_map = {
        'upper': 'Upper Body',
        'lower': 'Lower Body',
        'full': 'Full Body'
    }
    body_type = body_type_map.get(test_type, 'Upper Body')
    template_filename = f"{gym_folder} {body_type}.xlsm"
    return os.path.join(base_dir, gym_folder, template_filename)


def get_template_info(export_filename):
    """
    Determine which gym and body type based on the export filename.
    Returns: (gym_folder, template_filename)
    
    Expected filename patterns:
    - "motions" or "Motions" in name = Body Motions
    - "masters" or "Masters" in name = Body Masters
    
    - "upper" or "Upper" = Upper Body
    - "lower" or "Lower" = Lower Body
    - "full" or "Full" = Full Body
    """
    filename_lower = export_filename.lower()
    
    # Determine gym
    if "motions" in filename_lower:
        gym_folder = "Body Motions"
    elif "masters" in filename_lower:
        gym_folder = "Body Masters"
    else:
        # Default to Body Motions if not specified
        gym_folder = "Body Motions"
    
    # Determine body type
    if "upper" in filename_lower:
        body_type = "Upper Body"
    elif "lower" in filename_lower:
        body_type = "Lower Body"
    elif "full" in filename_lower:
        body_type = "Full Body"
    else:
        # Default to Upper Body if not specified
        body_type = "Upper Body"
    
    template_filename = f"{gym_folder} {body_type}.xlsm"
    
    return gym_folder, template_filename



def get_movement_label(movement, region):
    """
    Get the bilingual label for a movement/region combination.
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    movement_labels = {
        ('internal rotation', 'shoulder'): "Shoulder IR Standing Asymmetry / عدم توازن دوران الكتف الداخلي",
        ('external rotation', 'shoulder'): "Shoulder External Rotation Asymmetry /  عدم توازن الدوران الخارجي للكتف",
        ('flexion', 'shoulder'): "Shoulder Flexion Asymmetry /\n عدم تناسق انثناء الكتف",
        ('abduction', 'shoulder'): "Shoulder Abduction Asymmetry /\nعدم توازن إبعاد الكتف",
        ('push', 'shoulder'): "Shoulder Push Asymmetry/\nعدم توازن في دفع الكتف",
        ('pull', 'shoulder'): "Shoulder Pull Asymmetry/\nعدم توازن في سحب الكتف",
        ('extension', 'elbow'): "Elbow Extension Asymmetry /\nعدم توازن تمديد الكوع",
        ('flexion', 'elbow'): "Elbow Flexion Asymmetry /\nعدم توازن انثناء الكوع",
        ('grip squeeze', 'hand'): "Grip Squeeze Asymmetry/\nعدم توازن ضغط القبضة",
        ('extension', 'knee'): "Knee Extension Asymmetry /\n عدم تناسق تمديد الركبة",
        ('flexion', 'knee'): "Knee Flexion Asymmetry /\n عدم تناسق انثناء الركبة",
        ('abduction', 'hip'): "Hip Abduction Asymmetry /\n عدم تناسق إبعاد الورك",
        ('adduction', 'hip'): "Hip Adduction Asymmetry /\n عدم تناسق تقريب الورك",
        ('lateral flexion', 'trunk'): "Trunk Lateral Flexion /\nالثني الجانبي للجذع",
        ('flexion', 'hip'): "Hip Flexion Asymmetry /\nعدم تناسق ثني الورك",
        ('extension', 'hip'): "Hip Extension Asymmetry /\nعدم تناسق مدّ الورك",
    }
    
    return movement_labels.get((m, r), None)


def nz_str(v):
    if v is None:
        return ""
    return str(v)


def nz_float(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def calculate_trunk_asymmetry(patient_rows, src_ws):
    """
    Calculate trunk asymmetry from Lateral Flexion Right and Left.
    Returns: (percentage, weak_side) or (None, None) if not enough data.
    Formula: Asymmetry(%) = ((L - R) / ((L + R) / 2)) * 100
    """
    trunk_data = {}
    
    for row in patient_rows:
        movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
        region = nz_str(src_ws[f"H{row}"].value).lower().strip()
        
        # Look for trunk lateral flexion movements
        if region == "trunk" and "lateral flexion" in movement:
            force = nz_float(src_ws[f"R{row}"].value)  # N Avg Force (N)
            
            if "right" in movement:
                trunk_data['right'] = force
            elif "left" in movement:
                trunk_data['left'] = force
    
    # Need both right and left to calculate
    if 'right' not in trunk_data or 'left' not in trunk_data:
        return None, None
    
    right_force = trunk_data['right']
    left_force = trunk_data['left']
    
    # Skip if either is zero or invalid
    if right_force <= 0 or left_force <= 0:
        return None, None
    
    # Calculate asymmetry using the formula: ((L - R) / ((L + R) / 2)) * 100
    avg_force = (left_force + right_force) / 2
    asymmetry_pct = ((left_force - right_force) / avg_force) * 100
    
    # Determine weak side (smaller force value)
    if right_force < left_force:
        weak_side = "Right"
    else:
        weak_side = "Left"
    
    # Return absolute value of percentage
    return abs(asymmetry_pct), weak_side


def parse_asymmetry(raw):
    """
    raw example: '9.0% L' or '3% R' or 'n/a' or '0'
    returns: (pct_value, side_char) → (9.0, 'L')
    Returns (None, None) for invalid or n/a values
    """
    if raw is None:
        return None, None
    s = str(raw).strip().lower()
    if not s:
        return None, None
    
    # Check for n/a variants
    if s in ['n/a', 'na', 'n.a.', 'n.a']:
        return None, None

    # Check if it's just a number without % (like "0")
    if "%" not in s:
        # Try to parse as a plain number
        try:
            pct_val = float(s.replace(",", "."))
            # For plain numbers, assume no side character (will be handled as 0%)
            return pct_val, None
        except ValueError:
            return None, None

    percent_index = s.index("%")
    num_part = s[:percent_index].strip()   # e.g. "9.0"
    
    # Check if the numeric part is n/a
    if num_part.lower() in ['n/a', 'na', 'n.a.', 'n.a']:
        return None, None

    # side is last non-space character
    side_char = None
    for ch in reversed(s):
        if ch != " ":
            side_char = ch
            break

    try:
        pct_val = float(num_part.replace(",", "."))
    except ValueError:
        return None, None

    return pct_val, side_char


def get_target_info(movement, region):
    """
    Returns:
      (pct_cell_address, side_cell_address, right_text, left_text)
    or (None, None, None, None) if not relevant
    """
    m = movement.lower().strip()
    r = region.lower().strip()

    # ==================== UPPER BODY ====================
    
    # HAND – Grip Squeeze
    if r == "hand" and m == "grip squeeze":
        pct_cell = "AH21"
        side_cell = "AG22"
        right_text = "Right Grip Squeeze/\nضغط القبضة باليد اليمنى"
        left_text  = "Left Grip Squeeze/\nضغط القبضة باليد اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # ELBOW – Extension (Triceps)
    if r == "elbow" and m == "extension":
        pct_cell = "AB21"
        side_cell = "AA22"
        right_text = "Right Triceps /\n عضلات التراي سيبس اليمنى"
        left_text  = "Left Triceps /\n عضلات التراي سيبس اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # ELBOW – Flexion (Biceps)
    if r == "elbow" and m == "flexion":
        pct_cell = "AB23"
        side_cell = "AA24"
        right_text = "Right Biceps /\nعضلة الباي سيبس اليمنى"
        left_text  = "Left Biceps /\n عضلة الباي سيبس اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – Push
    if r == "shoulder" and m == "push":
        pct_cell = "P21"
        side_cell = "O22"
        right_text = "Right Shoulder Push/\nدفع الكتف الأيمن"
        left_text  = "Left Shoulder Push/\nدفع الكتف الأيسر"
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – Pull
    if r == "shoulder" and m == "pull":
        pct_cell = "P23"
        side_cell = "O24"
        right_text = "Right Shoulder Pull/\nسحب الكتف الأيمن"
        left_text  = "Left Shoulder Pull/\nسحب الكتف الأيسر"
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – External Rotation (ER)
    if r == "shoulder" and m == "external rotation":
        pct_cell = "D21"
        side_cell = "C22"
        right_text = "Right External Rotation /\nالدوران الخارجي الأيمن "
        left_text  = "Left External Rotation /\nالدوران الخارجي الأيسر "
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – Internal Rotation (IR)
    if r == "shoulder" and m == "internal rotation":
        pct_cell = "D23"
        side_cell = "C24"
        right_text = "Right Internal Rotation /\n الدوران الداخلي الأيمن"
        left_text  = "Left Internal Rotation / \nالدوران الداخلي الأيسر"
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – Flexion
    if r == "shoulder" and m == "flexion":
        pct_cell = "D25"
        side_cell = "C26"
        right_text = "Right shoulder flexion /\n ثني الكتف الأيمن"
        left_text  = "Left shoulder flexion /\n ثني الكتف الأيسر"
        return pct_cell, side_cell, right_text, left_text

    # SHOULDER – Abduction
    if r == "shoulder" and m == "abduction":
        pct_cell = "D27"
        side_cell = "C28"
        right_text = "Right shoulder abductor /\nعضلة فتح الكتف الأيمن"
        left_text  = "Left shoulder abductor /\nعضلة فتح الكتف الأيسر "
        return pct_cell, side_cell, right_text, left_text

    # ==================== LOWER BODY ====================
    # Lower body uses dynamic assignment, handled separately
    
    return None, None, None, None


def get_lower_body_cells(movement, region, knee_movements, hip_abd_add_movements, hip_flex_ext_movements):
    """
    Get cell addresses for lower body based on which movements are present.
    Returns: (label_cell, pct_cell, side_cell, right_text, left_text)
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    # KNEE movements - assigned to C21/C23 based on order
    if r == "knee":
        if (m, r) == ('extension', 'knee') and ('extension', 'knee') in knee_movements:
            idx = knee_movements.index(('extension', 'knee'))
            if idx == 0:
                return "C21", "D21", "C22", "Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى", "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"
            elif idx == 1:
                return "C23", "D23", "C24", "Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى", "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"
        elif (m, r) == ('flexion', 'knee') and ('flexion', 'knee') in knee_movements:
            idx = knee_movements.index(('flexion', 'knee'))
            if idx == 0:
                return "C21", "D21", "C22", "Right Hamstring /\nعضلات الفخذ الخلفية اليمنى", "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"
            elif idx == 1:
                return "C23", "D23", "C24", "Right Hamstring /\nعضلات الفخذ الخلفية اليمنى", "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"
    
    # HIP abd/add - assigned to O21/O23 based on order
    if r == "hip" and m in ['adduction', 'abduction']:
        if (m, r) == ('adduction', 'hip') and ('adduction', 'hip') in hip_abd_add_movements:
            idx = hip_abd_add_movements.index(('adduction', 'hip'))
            if idx == 0:
                return "O21", "P21", "O22", "Right Adductors /\nعضلات الفخذ الداخلي اليمنى", "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"
            elif idx == 1:
                return "O23", "P23", "O24", "Right Adductors /\nعضلات الفخذ الداخلي اليمنى", "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"
        elif (m, r) == ('abduction', 'hip') and ('abduction', 'hip') in hip_abd_add_movements:
            idx = hip_abd_add_movements.index(('abduction', 'hip'))
            if idx == 0:
                return "O21", "P21", "O22", "Right Abductors /\nعضلات الفخذ الخارجية اليمنى", "Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"
            elif idx == 1:
                return "O23", "P23", "O24", "Right Abductors /\nعضلات الفخذ الخارجية اليمنى", "Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"
    
    # HIP flex/ext - assigned to AG21/AG23 based on order
    if r == "hip" and m in ['flexion', 'extension']:
        if (m, r) == ('flexion', 'hip') and ('flexion', 'hip') in hip_flex_ext_movements:
            idx = hip_flex_ext_movements.index(('flexion', 'hip'))
            if idx == 0:
                return "AG21", "AH21", "AG22", "Right Hip Flexors /\nعضلات مثنية الورك اليمنى", "Left Hip Flexors /\nعضلات مثنية الورك اليسرى"
            elif idx == 1:
                return "AG23", "AH23", "AG24", "Right Hip Flexors /\nعضلات مثنية الورك اليمنى", "Left Hip Flexors /\nعضلات مثنية الورك اليسرى"
        elif (m, r) == ('extension', 'hip') and ('extension', 'hip') in hip_flex_ext_movements:
            idx = hip_flex_ext_movements.index(('extension', 'hip'))
            if idx == 0:
                return "AG21", "AH21", "AG22", "Right Hip Extensors /\nعضلات باسطة الورك اليمنى", "Left Hip Extensors /\nعضلات باسطة الورك اليسرى"
            elif idx == 1:
                return "AG23", "AH23", "AG24", "Right Hip Extensors /\nعضلات باسطة الورك اليمنى", "Left Hip Extensors /\nعضلات باسطة الورك اليسرى"
    
    return None, None, None, None, None


def get_upper_body_cells(movement, region, shoulder_c_movements, shoulder_o_movements, elbow_movements, hand_movements):
    """
    Get cell addresses for upper body based on which movements are present.
    Returns: (label_cell, pct_cell, side_cell, right_text, left_text)
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    # SHOULDER rotations/movements - C21/C23/C25/C27
    if r == "shoulder" and m in ['external rotation', 'internal rotation', 'flexion', 'abduction']:
        if (m, r) in shoulder_c_movements:
            idx = shoulder_c_movements.index((m, r))
            cells = [
                ("C21", "D21", "C22"),  # idx 0
                ("C23", "D23", "C24"),  # idx 1
                ("C25", "D25", "C26"),  # idx 2
                ("C27", "D27", "C28"),  # idx 3
            ]
            if idx < len(cells):
                label_cell, pct_cell, side_cell = cells[idx]
                if m == 'external rotation':
                    return label_cell, pct_cell, side_cell, "Right External Rotation /\nالدوران الخارجي الأيمن ", "Left External Rotation /\nالدوران الخارجي الأيسر "
                elif m == 'internal rotation':
                    return label_cell, pct_cell, side_cell, "Right Internal Rotation /\n الدوران الداخلي الأيمن", "Left Internal Rotation / \nالدوران الداخلي الأيسر"
                elif m == 'flexion':
                    return label_cell, pct_cell, side_cell, "Right shoulder flexion /\n ثني الكتف الأيمن", "Left shoulder flexion /\n ثني الكتف الأيسر"
                elif m == 'abduction':
                    return label_cell, pct_cell, side_cell, "Right shoulder abductor /\nعضلة فتح الكتف الأيمن", "Left shoulder abductor /\nعضلة فتح الكتف الأيسر "
    
    # SHOULDER push/pull - O21/O23
    if r == "shoulder" and m in ['push', 'pull']:
        if (m, r) in shoulder_o_movements:
            idx = shoulder_o_movements.index((m, r))
            if idx == 0:
                label_cell, pct_cell, side_cell = "O21", "P21", "O22"
            elif idx == 1:
                label_cell, pct_cell, side_cell = "O23", "P23", "O24"
            else:
                return None, None, None, None, None
            
            if m == 'push':
                return label_cell, pct_cell, side_cell, "Right Shoulder Push/\nدفع الكتف الأيمن", "Left Shoulder Push/\nدفع الكتف الأيسر"
            elif m == 'pull':
                return label_cell, pct_cell, side_cell, "Right Shoulder Pull/\nسحب الكتف الأيمن", "Left Shoulder Pull/\nسحب الكتف الأيسر"
    
    # ELBOW - AA21/AA23
    if r == "elbow" and m in ['extension', 'flexion']:
        if (m, r) in elbow_movements:
            idx = elbow_movements.index((m, r))
            if idx == 0:
                label_cell, pct_cell, side_cell = "AA21", "AB21", "AA22"
            elif idx == 1:
                label_cell, pct_cell, side_cell = "AA23", "AB23", "AA24"
            else:
                return None, None, None, None, None
            
            if m == 'extension':
                return label_cell, pct_cell, side_cell, "Right Triceps /\n عضلات التراي سيبس اليمنى", "Left Triceps /\n عضلات التراي سيبس اليسرى"
            elif m == 'flexion':
                return label_cell, pct_cell, side_cell, "Right Biceps /\nعضلة الباي سيبس اليمنى", "Left Biceps /\n عضلة الباي سيبس اليسرى"
    
    # HAND - AG21 only
    if r == "hand" and m == 'grip squeeze':
        if (m, r) in hand_movements:
            return "AG21", "AH21", "AG22", "Right Grip Squeeze/\nضغط القبضة باليد اليمنى", "Left Grip Squeeze/\nضغط القبضة باليد اليسرى"
    
    return None, None, None, None, None


def get_full_body_cells(movement, region, shoulder_c_movements, elbow_o_movements, knee_aa_movements, hip_ag_movements):
    """
    Get cell addresses for full body based on which movements are present.
    Returns: (label_cell, pct_cell, side_cell, right_text, left_text)
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    # SHOULDER rotations/movements - C21/C23/C25/C27
    if r == "shoulder" and m in ['external rotation', 'internal rotation', 'flexion', 'abduction']:
        if (m, r) in shoulder_c_movements:
            idx = shoulder_c_movements.index((m, r))
            cells = [
                ("C21", "D21", "C22"),  # idx 0
                ("C23", "D23", "C24"),  # idx 1
                ("C25", "D25", "C26"),  # idx 2
                ("C27", "D27", "C28"),  # idx 3
            ]
            if idx < len(cells):
                label_cell, pct_cell, side_cell = cells[idx]
                if m == 'external rotation':
                    return label_cell, pct_cell, side_cell, "Right External Rotation /\nالدوران الخارجي الأيمن ", "Left External Rotation /\nالدوران الخارجي الأيسر "
                elif m == 'internal rotation':
                    return label_cell, pct_cell, side_cell, "Right Internal Rotation /\n الدوران الداخلي الأيمن", "Left Internal Rotation / \nالدوران الداخلي الأيسر"
                elif m == 'flexion':
                    return label_cell, pct_cell, side_cell, "Right shoulder flexion /\n ثني الكتف الأيمن", "Left shoulder flexion /\n ثني الكتف الأيسر"
                elif m == 'abduction':
                    return label_cell, pct_cell, side_cell, "Right shoulder abductor /\nعضلة فتح الكتف الأيمن", "Left shoulder abductor /\nعضلة فتح الكتف الأيسر "
    
    # ELBOW - O21/O23
    if r == "elbow" and m in ['extension', 'flexion']:
        if (m, r) in elbow_o_movements:
            idx = elbow_o_movements.index((m, r))
            if idx == 0:
                label_cell, pct_cell, side_cell = "O21", "P21", "O22"
            elif idx == 1:
                label_cell, pct_cell, side_cell = "O23", "P23", "O24"
            else:
                return None, None, None, None, None
            
            if m == 'extension':
                return label_cell, pct_cell, side_cell, "Right Triceps /\n عضلات التراي سيبس اليمنى", "Left Triceps /\n عضلات التراي سيبس اليسرى"
            elif m == 'flexion':
                return label_cell, pct_cell, side_cell, "Right Biceps /\nعضلة الباي سيبس اليمنى", "Left Biceps /\n عضلة الباي سيبس اليسرى"
    
    # KNEE - AA21/AA23
    if r == "knee" and m in ['extension', 'flexion']:
        if (m, r) in knee_aa_movements:
            idx = knee_aa_movements.index((m, r))
            if idx == 0:
                label_cell, pct_cell, side_cell = "AA21", "AB21", "AA22"
            elif idx == 1:
                label_cell, pct_cell, side_cell = "AA23", "AB23", "AA24"
            else:
                return None, None, None, None, None
            
            if m == 'extension':
                return label_cell, pct_cell, side_cell, "Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى", "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"
            elif m == 'flexion':
                return label_cell, pct_cell, side_cell, "Right Hamstring /\nعضلات الفخذ الخلفية اليمنى", "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"
    
    # HIP abd/add - AG21/AG23
    if r == "hip" and m in ['abduction', 'adduction']:
        if (m, r) in hip_ag_movements:
            idx = hip_ag_movements.index((m, r))
            if idx == 0:
                label_cell, pct_cell, side_cell = "AG21", "AH21", "AG22"
            elif idx == 1:
                label_cell, pct_cell, side_cell = "AG23", "AH23", "AG24"
            else:
                return None, None, None, None, None
            
            if m == 'abduction':
                return label_cell, pct_cell, side_cell, "Right Abductors /\nعضلات الفخذ الخارجية اليمنى", "Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"
            elif m == 'adduction':
                return label_cell, pct_cell, side_cell, "Right Adductors /\nعضلات الفخذ الداخلي اليمنى", "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"
    
    return None, None, None, None, None


def get_target_info_full_body(movement, region):
    """
    Special handling for Full Body tests where some movements map to different cells.
    Returns same format as get_target_info.
    """
    m = movement.lower().strip()
    r = region.lower().strip()
    
    # ELBOW – Extension (Triceps) - Full Body position
    if r == "elbow" and m == "extension":
        pct_cell = "P21"
        side_cell = "O22"
        right_text = "Right Triceps /\n عضلات التراي سيبس اليمنى"
        left_text  = "Left Triceps /\n عضلات التراي سيبس اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # ELBOW – Flexion (Biceps) - Full Body position
    if r == "elbow" and m == "flexion":
        pct_cell = "P23"
        side_cell = "O24"
        right_text = "Right Biceps /\nعضلة الباي سيبس اليمنى"
        left_text  = "Left Biceps /\n عضلة الباي سيبس اليسرى"
        return pct_cell, side_cell, right_text, left_text
    
    # KNEE – Extension (Quadriceps) - Full Body position
    if r == "knee" and m == "extension":
        pct_cell = "AB21"
        side_cell = "AA22"
        right_text = "Right Quadriceps /\nعضلات الفخذ الأمامية اليمنى"
        left_text  = "Left Quadriceps /\nعضلات الفخذ الأمامية اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # KNEE – Flexion (Hamstring) - Full Body position
    if r == "knee" and m == "flexion":
        pct_cell = "AB23"
        side_cell = "AA24"
        right_text = "Right Hamstring /\nعضلات الفخذ الخلفية اليمنى"
        left_text  = "Left Hamstring /\nعضلات الفخذ الخلفية اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # HIP – Abduction - Full Body position
    if r == "hip" and m == "abduction":
        pct_cell = "AH21"
        side_cell = "AG22"
        right_text = "Right Abductors /\nعضلات الفخذ الخارجية اليمنى"
        left_text  = "Left Abductors /\nعضلات الفخذ  الخارجية اليسرى"
        return pct_cell, side_cell, right_text, left_text

    # HIP – Adduction - Full Body position
    if r == "hip" and m == "adduction":
        pct_cell = "AH23"
        side_cell = "AG24"
        right_text = "Right Adductors /\nعضلات الفخذ الداخلي اليمنى"
        left_text  = "Left Adductors /\nعضلات الفخذ الداخلي اليسرى"
        return pct_cell, side_cell, right_text, left_text
    
    # For all other movements in full body, use standard mapping
    return get_target_info(movement, region)


def clear_fields(ws):
    """
    Clears only the cells we are controlling.
    """
    addresses = [
        "AH21", "AH22", "AH23", "AH24",
        "AB21", "AB22", "AB23", "AB24",
        "P21", "P22", "P23", "P24",
        "D21", "C22", "D23", "C24", "D25", "C26", "D27", "C28",
        "AA21", "AA22", "AA23", "AA24",
        "AG21", "AG22", "AG23", "AG24",
        "A6", "A21"
    ]
    for addr in addresses:
        ws[addr].value = None


def make_safe_filename(name):
    """
    Remove characters that are illegal in file names and trim.
    """
    invalid = '<>:"/\\|?*'
    safe = "".join("_" if c in invalid else c for c in name)
    safe = safe.strip()
    if not safe:
        safe = "Report"
    return safe + ".xlsm"


def get_remark_for_percentage(pct_value):
    """
    Return bilingual remark based on asymmetry percentage.
    """
    if pct_value is None:
        return ""
    
    pct = abs(pct_value) * 100  # Convert to percentage
    
    if 0.1 <= pct <= 3.9:
        return "Perfect Symmetry / \nتناظر مثالي"
    elif 4 <= pct <= 7.9:
        return "Normal Symmetry / \nتناظر طبيعي"
    elif 8 <= pct <= 14.9:
        return "Weakness / \nضعف"
    elif 15 <= pct <= 19.9:
        return "Problem / \nمشكلة"
    elif 20 <= pct <= 29.9:
        return "Major Problem / \nمشكلة كبيرة"
    elif pct >= 30:
        return "Risk Of Injury / \nخطر الإصابة"
    else:
        return ""


def load_test_log(gym_folder, base_dir):
    """
    Load the test log for the specified gym.
    Returns a dict: {patient_name: {test_type: {date: True}}}
    """
    log_file = os.path.join(base_dir, f"{gym_folder}_test_log.json")
    if os.path.exists(log_file):
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}


def save_test_log(gym_folder, base_dir, log_data):
    """
    Save the test log for the specified gym.
    """
    log_file = os.path.join(base_dir, f"{gym_folder}_test_log.json")
    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)


def log_test(gym_folder, base_dir, patient_name, test_type, test_date, movement_count=0):
    """
    Log a test to the gym's test log with movement count.
    """
    log_data = load_test_log(gym_folder, base_dir)
    
    # Normalize patient name
    patient_name = re.sub(r'\s+', ' ', patient_name).strip()
    
    # Convert date to string if needed
    if isinstance(test_date, datetime):
        date_str = test_date.strftime('%Y-%m-%d')
    else:
        date_str = str(test_date)
    
    # Create structure if needed
    if patient_name not in log_data:
        log_data[patient_name] = {}
    if test_type not in log_data[patient_name]:
        log_data[patient_name][test_type] = {}
    
    # Log this date with movement count
    log_data[patient_name][test_type][date_str] = movement_count
    
    save_test_log(gym_folder, base_dir, log_data)


def check_for_new_tests(export_path, gym_folder, base_dir):
    """
    Check export file against existing logs and report new tests or updated tests.
    Returns a list of new/updated tests found.
    Uses the same test type detection and movement filtering logic as normal processing.
    """
    log_data = load_test_log(gym_folder, base_dir)
    new_tests = []
    
    # Load export file
    wb = load_workbook(export_path, data_only=True)
    src_ws = wb.active
    
    # Collect rows per patient name (same as normal processing)
    patients_rows = {}
    for row in range(2, src_ws.max_row + 1):
        name_val = nz_str(src_ws[f"A{row}"].value).strip()
        if not name_val:
            continue
        # Normalize patient name - remove extra spaces (same as normal processing)
        name_val = re.sub(r'\s+', ' ', name_val).strip()
        if name_val not in patients_rows:
            patients_rows[name_val] = []
        patients_rows[name_val].append(row)
    
    # Process each patient using same logic as normal processing
    patients_tests = {}  # patient_name -> {test_type: {date: movement_count}}
    
    for patient_name, rows in patients_rows.items():
        # Use detect_test_type to determine test types (same as normal processing)
        test_types = detect_test_type(rows, src_ws)
        if isinstance(test_types, str):
            test_types = [test_types]
        
        patients_tests[patient_name] = {}
        
        # Process each test type for this patient
        for test_type in test_types:
            # Collect movements with valid asymmetry data (same filtering as normal processing)
            movements_present = {}
            movements_stored = {}  # Track which movements would actually be stored (highest asymmetry only)
            
            for row in rows:
                movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
                region = nz_str(src_ws[f"H{row}"].value).lower().strip()
                asym_raw = src_ws[f"S{row}"].value
                
                # Skip if no movement/region
                if not movement or not region:
                    continue
                
                # Filter by test type if multiple test types exist
                if len(test_types) > 1:
                    row_test_type = get_movement_test_type(movement, region)
                    if row_test_type and row_test_type != test_type:
                        continue
                
                # Skip trunk for normal movements (it's handled separately in lower body)
                if test_type == 'lower' and region == "trunk":
                    # Check if trunk will be calculated
                    trunk_pct, _ = calculate_trunk_asymmetry(rows, src_ws)
                    if trunk_pct is not None:
                        # Trunk will be included, add it once
                        if ('lateral flexion', 'trunk') not in movements_stored:
                            movements_stored[('lateral flexion', 'trunk')] = trunk_pct
                    continue
                
                # Skip if no asymmetry data
                if asym_raw in (None, ""):
                    continue
                
                # Only include if valid asymmetry can be parsed
                pct_value, side_char = parse_asymmetry(asym_raw)
                if pct_value is None:
                    continue
                
                # Track this movement with its asymmetry value
                key = (movement, region)
                if key not in movements_present:
                    movements_present[key] = []
                movements_present[key].append(abs(pct_value))
            
            # Now determine which movements would actually be stored (only largest asymmetry per movement)
            for key, asymmetry_values in movements_present.items():
                # Only store the movement with the largest asymmetry (same logic as normal processing)
                movements_stored[key] = max(asymmetry_values)
            
            # Get date from first row
            date_val = None
            for row in rows:
                date_val = src_ws[f"C{row}"].value
                if date_val:
                    break
            
            if date_val:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val)
                
                # Count only movements that would actually be stored (one per unique movement/region pair)
                movement_count = len(movements_stored)
                
                if test_type not in patients_tests[patient_name]:
                    patients_tests[patient_name][test_type] = {}
                
                patients_tests[patient_name][test_type][date_str] = movement_count
    
    wb.close()
    
    # Compare against log
    for patient_name, test_types_data in patients_tests.items():
        for test_type, dates in test_types_data.items():
            for date_str, movement_count in dates.items():
                # Check if this test exists in log
                is_new = True
                is_updated = False
                old_count = 0
                
                if patient_name in log_data:
                    if test_type in log_data[patient_name]:
                        if date_str in log_data[patient_name][test_type]:
                            old_count = log_data[patient_name][test_type][date_str]
                            is_new = False
                            # Check if movement count increased
                            if movement_count > old_count:
                                is_updated = True
                
                if is_new or is_updated:
                    new_tests.append({
                        'patient': patient_name,
                        'test_type': test_type,
                        'date': date_str,
                        'movement_count': movement_count,
                        'old_count': old_count,
                        'status': 'NEW' if is_new else 'UPDATED'
                    })
    
    return new_tests


def fill_template_with_xlwings(template_path, out_path, patient_name, patient_data, gym_folder):
    """
    Use xlwings to fill data while preserving all Excel features like data validation.
    """
    app = None
    wb = None
    try:
        # Copy the template to the output location first
        shutil.copy2(template_path, out_path)
        
        # Open Excel invisibly with all alerts/prompts disabled
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        app.api.AskToUpdateLinks = False
        app.api.DisplayAlerts = False
        app.api.AlertBeforeOverwriting = False
        
        # Open the copied file
        wb = app.books.open(out_path, update_links=False, read_only=False, ignore_read_only_recommended=True)
        workbook_name = wb.name
        
        # Determine sheet name based on gym folder
        template_sheet = gym_folder  # "Body Masters" or "Body Motions"
        
        # Find the correct sheet
        ws = None
        for sheet in wb.sheets:
            if sheet.name == template_sheet:
                ws = sheet
                break
        
        if ws is None:
            available_sheets = [s.name for s in wb.sheets]
            wb.close()
            app.quit()
            raise Exception(f"Sheet '{template_sheet}' not found. Available sheets: {available_sheets}")
        
        # Disable events during data entry
        app.api.EnableEvents = False
        
        # Normalize patient name - remove extra spaces
        normalized_name = re.sub(r'\s+', ' ', patient_name).strip()
        
        # Set name in A6
        ws.range('A6').value = normalized_name
        
        # Set date in A21
        if patient_data.get('date'):
            ws.range('A21').value = patient_data['date']
        
        # Set all other cells
        for cell_addr, cell_value in patient_data.get('cells', {}).items():
            ws.range(cell_addr).value = cell_value
        
        # Add remarks for all percentage cells
        percentage_cells = [
            ('D21', 'D22'), ('D23', 'D24'), ('D25', 'D26'), ('D27', 'D28'),
            ('P21', 'P22'), ('P23', 'P24'), ('P25', 'P26'), ('P27', 'P28'),
            ('AB21', 'AB22'), ('AB23', 'AB24'), ('AB25', 'AB26'), ('AB27', 'AB28'),
            ('AH21', 'AH22'), ('AH23', 'AH24'), ('AH25', 'AH26'), ('AH27', 'AH28')
        ]
        
        for pct_cell, remark_cell in percentage_cells:
            pct_value = ws.range(pct_cell).value
            if pct_value is not None and pct_value != "":
                remark = get_remark_for_percentage(pct_value)
                if remark:
                    ws.range(remark_cell).value = remark
        
        # Determine which body parts are present and populate A11-A17 (or A11-A18)
        body_parts_present = set()
        test_type = patient_data.get('test_type', 'upper')
        movements_stored = patient_data.get('movements', [])
        
        if test_type == 'lower':
            # LOWER BODY - check which movements were actually stored
            for movement, region in movements_stored:
                if region == 'knee':
                    if movement == 'extension':
                        body_parts_present.add('quadriceps')
                    elif movement == 'flexion':
                        body_parts_present.add('hamstring')
                elif region == 'hip':
                    if movement == 'adduction':
                        body_parts_present.add('hip_adductors')
                    elif movement == 'abduction':
                        body_parts_present.add('hip_abductors')
                    elif movement == 'flexion':
                        body_parts_present.add('hip_flexors')
                    elif movement == 'extension':
                        body_parts_present.add('hip_extensors')
                elif region == 'trunk':
                    body_parts_present.add('trunk')
                
        elif test_type == 'upper':
            # UPPER BODY - check which movements were actually stored
            for movement, region in movements_stored:
                if region == 'shoulder':
                    if movement in ['external rotation', 'internal rotation']:
                        body_parts_present.add('rotator_cuff')
                    elif movement in ['flexion', 'abduction']:
                        body_parts_present.add('shoulder')
                    elif movement == 'push':
                        body_parts_present.add('chest')
                    elif movement == 'pull':
                        body_parts_present.add('back')
                elif region == 'elbow':
                    if movement == 'extension':
                        body_parts_present.add('triceps')
                    elif movement == 'flexion':
                        body_parts_present.add('biceps')
                elif region == 'hand':
                    if movement == 'grip squeeze':
                        body_parts_present.add('hand')
                
        elif test_type == 'full':
            # FULL BODY - check which movements were actually stored
            for movement, region in movements_stored:
                if region == 'shoulder':
                    if movement in ['external rotation', 'internal rotation']:
                        body_parts_present.add('rotator_cuff')
                    elif movement in ['flexion', 'abduction']:
                        body_parts_present.add('shoulder')
                elif region == 'elbow':
                    if movement == 'extension':
                        body_parts_present.add('triceps')
                    elif movement == 'flexion':
                        body_parts_present.add('biceps')
                elif region == 'knee':
                    if movement == 'extension':
                        body_parts_present.add('quadriceps')
                    elif movement == 'flexion':
                        body_parts_present.add('hamstring')
                elif region == 'hip':
                    if movement == 'abduction':
                        body_parts_present.add('hip_abductors')
                    elif movement == 'adduction':
                        body_parts_present.add('hip_adductors')
        
        # Map body parts to their bilingual text
        body_part_map = {
            # Upper Body
            'shoulder': "Shoulder / الكتف",
            'rotator_cuff': "Rotator cuff/ الكُمّ",
            'chest': "Chest/ الصدر",
            'back': "Back/ الظهر",
            'triceps': "Triceps /التراي سيبس",
            'biceps': "Biceps /الباي سيبس",
            'hand': "Hand/ اليد",
            # Lower Body
            'quadriceps': "Quadriceps / الفخذ الأمامي",
            'hamstring': "Hamstring / الفخذ الخلفي",
            'hip_abductors': "Hip Abductors / الفخذ  الخارجي",
            'hip_adductors': "Hip Adductors / الفخذ الداخلي",
            'hip_flexors': "Hip Flexors / عضلات ثني الورك",
            'hip_extensors': "Hip Extensors/عضلات بسط الورك",
            'trunk': "Trunk / الجذع"
        }
        
        # Fill A11-A18 with the body parts present
        body_parts_list = []
        
        # Use test_type from patient_data to determine correct order
        test_type = patient_data.get('test_type', 'upper')
        
        if test_type == 'upper':
            # Order for upper body
            ordered_parts = ['shoulder', 'rotator_cuff', 'chest', 'back', 'triceps', 'biceps', 'hand']
            max_slots = 7
        elif test_type == 'lower':
            # Order for lower body
            ordered_parts = ['quadriceps', 'hamstring', 'hip_abductors', 'hip_adductors', 'hip_flexors', 'hip_extensors', 'trunk']
            max_slots = 8
        else:  # full body
            # Order for full body
            ordered_parts = ['shoulder', 'rotator_cuff', 'triceps', 'biceps', 'hamstring', 'quadriceps', 'hip_abductors', 'hip_adductors']
            max_slots = 8
        
        for part in ordered_parts:
            if part in body_parts_present:
                body_parts_list.append(body_part_map[part])
        
        # Write to cells A11-A17 or A11-A18
        for i, body_part_text in enumerate(body_parts_list):
            if i < max_slots:
                ws.range(f'A{11+i}').value = body_part_text
        
        # Save the Excel file
        wb.save()
        
        # Export to PDF using AppleScript while workbook is still open
        pdf_path = out_path.replace('.xlsm', '.pdf')
        try:
            applescript = f'''
            tell application "Microsoft Excel"
                set visible to false
                if not (exists workbook "{workbook_name}") then
                    open "{out_path}"
                end if
                tell workbook "{workbook_name}"
                    save active sheet in POSIX file "{pdf_path}" as PDF file format
                end tell
            end tell
            '''
            subprocess.run(['osascript', '-e', applescript], capture_output=True, timeout=30, text=True)
        except Exception:
            pass
        
        # Close workbook and quit Excel
        wb.close()
        app.quit()
        app.quit()
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        if wb:
            try:
                wb.close()
            except:
                pass
        if app:
            try:
                app.quit()
            except:
                pass
        return False


def main():
    if len(sys.argv) < 2:
        print("Usage: process_dynamo.py <export_file.xlsx>")
        print("   or: process_dynamo.py --report <masters|motions>")
        sys.exit(1)

    # Check for report-only mode
    if sys.argv[1] == "--report":
        if len(sys.argv) < 3 or sys.argv[2].lower() not in ['masters', 'motions', 'all']:
            print("Usage: process_dynamo.py --report <masters|motions|all>")
            sys.exit(1)
        
        gym_choice = sys.argv[2].lower()
        base_dir = os.path.dirname(os.path.abspath(__file__))
        
        if gym_choice == 'all':
            # Generate combined report for both gyms
            gyms = ['Body Masters', 'Body Motions']
            report_path = os.path.join(base_dir, f"Combined_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
            
            # Calculate total across both gyms
            total_all = 0
            total_patients_all = 0
            for gym in gyms:
                log_data = load_test_log(gym, base_dir)
                total_all += sum(
                    len(dates)
                    for patient in log_data.values()
                    for test_type in patient.values()
                    for dates in [test_type]
                )
                total_patients_all += len(log_data)
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"Combined Test Database Summary\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("="*60 + "\n")
                f.write(f"Total tests (both gyms): {total_all}\n")
                f.write(f"Total patients (both gyms): {total_patients_all}\n")
                f.write("="*60 + "\n\n")
                
                for gym_folder in gyms:
                    log_data = load_test_log(gym_folder, base_dir)
                    
                    if not log_data:
                        f.write(f"\n{gym_folder}\n")
                        f.write("-"*60 + "\n")
                        f.write("No data found\n\n")
                        continue
                    
                    total_count = sum(
                        len(dates) 
                        for patient in log_data.values() 
                        for test_type in patient.values() 
                        for dates in [test_type]
                    )
                    
                    f.write(f"\n{gym_folder}\n")
                    f.write("-"*60 + "\n")
                    f.write(f"Total tests: {total_count}\n")
                    f.write(f"Total patients: {len(log_data)}\n")
                    f.write("-"*60 + "\n")
                    
                    for patient_name in sorted(log_data.keys()):
                        patient_tests = log_data[patient_name]
                        patient_total = sum(len(dates) for dates in patient_tests.values())
                        f.write(f"\n{patient_name} ({patient_total} test(s))\n")
                        for test_type, dates in patient_tests.items():
                            f.write(f"  {test_type.capitalize()} Body: {len(dates)} test(s)\n")
                            for date_str, movement_count in sorted(dates.items()):
                                f.write(f"    - {date_str} ({movement_count} movements)\n")
                    
                    f.write("\n")
            
            print(f"Summary report saved: {report_path}")
            total_all = sum(
                len(dates)
                for gym in gyms
                for patient in load_test_log(gym, base_dir).values()
                for test_type in patient.values()
                for dates in [test_type]
            )
            print(f"Total tests across both gyms: {total_all}")
            sys.exit(0)
        
        gym_folder = "Body Masters" if gym_choice == "masters" else "Body Motions"
        base_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Load log data
        log_data = load_test_log(gym_folder, base_dir)
        
        if not log_data:
            print(f"No data found for {gym_folder}")
            sys.exit(0)
        
        # Count totals
        total_count = sum(
            len(dates) 
            for patient in log_data.values() 
            for test_type in patient.values() 
            for dates in [test_type]
        )
        
        # Create summary report
        report_path = os.path.join(base_dir, f"{gym_folder}_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"Test Database Summary - {gym_folder}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n")
            f.write(f"Total tests in database: {total_count}\n")
            f.write(f"Total patients: {len(log_data)}\n")
            f.write("="*60 + "\n\n")
            
            # List all patients with their test counts
            for patient_name in sorted(log_data.keys()):
                patient_tests = log_data[patient_name]
                patient_total = sum(len(dates) for dates in patient_tests.values())
                f.write(f"\n{patient_name} ({patient_total} test(s))\n")
                for test_type, dates in patient_tests.items():
                    f.write(f"  {test_type.capitalize()} Body: {len(dates)} test(s)\n")
                    for date_str, movement_count in sorted(dates.items()):
                        f.write(f"    - {date_str} ({movement_count} movements)\n")
        
        print(f"Summary report saved: {report_path}")
        print(f"Total tests in database: {total_count}")
        print(f"Total patients: {len(log_data)}")
        sys.exit(0)

    export_path = sys.argv[1]
    if not os.path.isfile(export_path):
        print(f"File not found: {export_path}")
        sys.exit(1)
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Determine which template to use based on export filename
    export_filename = os.path.basename(export_path)
    gym_folder, template_filename = get_template_info(export_filename)
    
    # Check if this is a "check" mode (filename contains "check")
    is_check_mode = "check" in export_filename.lower()
    
    if is_check_mode:
        print(f"CHECK MODE: Comparing against existing logs for {gym_folder}...")
        new_tests = check_for_new_tests(export_path, gym_folder, base_dir)
        
        # Create output report
        desktop_path = os.path.expanduser("~/Desktop")
        report_path = os.path.join(desktop_path, f"{gym_folder}_new_tests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        # Count total tests in log (don't modify log during check)
        log_data = load_test_log(gym_folder, base_dir)
        total_count = sum(
            len(dates) 
            for patient in log_data.values() 
            for test_type in patient.values() 
            for dates in [test_type]
        )
        
        # Count new vs updated
        new_count = sum(1 for test in new_tests if test['status'] == 'NEW')
        updated_count = sum(1 for test in new_tests if test['status'] == 'UPDATED')
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(f"New Tests Report - {gym_folder}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*60 + "\n")
            f.write(f"Total tests in database: {total_count}\n")
            f.write(f"New tests: {new_count}\n")
            f.write(f"Updated tests: {updated_count}\n")
            f.write("="*60 + "\n\n")
            
            if new_tests:
                f.write(f"New/Updated test details:\n\n")
                for test in new_tests:
                    f.write(f"Status: {test['status']}\n")
                    f.write(f"Patient: {test['patient']}\n")
                    f.write(f"Test Type: {test['test_type'].capitalize()} Body\n")
                    f.write(f"Test Date: {test['date']}\n")
                    f.write(f"Movements: {test['movement_count']}")
                    if test['status'] == 'UPDATED':
                        f.write(f" (was {test['old_count']})")
                    f.write("\n")
                    f.write("-"*40 + "\n")
            else:
                f.write("No new tests found. All tests in export have been processed before.\n")
        
        print(f"\nReport saved: {report_path}")
        print(f"Total tests in database: {total_count}")
        print(f"Found {len(new_tests)} new/updated test(s)")
        if new_tests:
            print(f"Note: Run these tests normally (without 'check' in filename) to add them to the database.")
        sys.exit(0)
    
    # Normal processing mode
    # Request automation permissions upfront to avoid repeated prompts
    try:
        subprocess.run([
            'osascript', '-e',
            'tell application "System Events" to get name of every process'
        ], capture_output=True, timeout=5)
    except:
        pass  # If this fails, xlwings will handle permissions itself

    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Determine which template to use based on export filename
    export_filename = os.path.basename(export_path)
    gym_folder, template_filename = get_template_info(export_filename)
    
    # Determine body type for proper cell mapping
    is_full_body = "full" in export_filename.lower()
    
    template_path = os.path.join(base_dir, gym_folder, template_filename)
    
    print(f"Using template: {gym_folder}/{template_filename}")

    if not os.path.isfile(template_path):
        print(f"Template not found: {template_path}")
        print(f"Make sure the file exists in the '{gym_folder}' folder.")
        sys.exit(1)

    # Load export workbook once
    src_wb = load_workbook(export_path, data_only=True)
    src_ws = src_wb.active
    max_row = src_ws.max_row

    # 1) Collect rows per patient name
    patients = {}  # name -> list of row indices
    for row in range(2, max_row + 1):  # assuming row 1 is header
        name_val = nz_str(src_ws[f"A{row}"].value).strip()
        if not name_val:
            continue
        if name_val not in patients:
            patients[name_val] = []
        patients[name_val].append(row)

    if not patients:
        print("No names found in column A. Nothing to do.")
        sys.exit(0)

    num_patients = len(patients)
    export_base = os.path.splitext(os.path.basename(export_path))[0]

    # 2) Create main Programs folder on Desktop
    desktop_path = os.path.expanduser("~/Desktop")
    programs_folder = os.path.join(desktop_path, "Programs")
    os.makedirs(programs_folder, exist_ok=True)

    print(f"Found {num_patients} patient(s).")

    # 3) Process each patient separately
    for patient_name, rows in patients.items():
        print(f"Processing: {patient_name} ({len(rows)} row(s))")

        # Detect which test type(s) this patient has
        test_types = detect_test_type(rows, src_ws)
        if isinstance(test_types, str):
            test_types = [test_types]
        
        print(f"  Detected test type(s): {', '.join(test_types)}")
        
        # Process each test type for this patient
        for test_type in test_types:
            print(f"  Creating {test_type} body report...")
            
            # Determine if this is full body for proper cell mapping
            is_full_body = (test_type == 'full')
            
            # Get the appropriate template
            template_path = get_template_for_test_type(test_type, gym_folder, base_dir)
            
            if not os.path.isfile(template_path):
                print(f"    Template not found: {template_path}")
                continue
            
            # Collect all data for this patient and test type
            patient_data = {
                'date': None,
                'cells': {},
                'test_type': test_type,  # Store test type for body parts detection
                'movements': []  # Track which movements were actually stored
            }
            
            date_set = False
            
            # Collect movements present for dynamic cell assignment
            # Only include movements that have valid asymmetry percentages
            movements_present = {}
            if test_type in ['lower', 'upper', 'full']:
                for row in rows:
                    movement = nz_str(src_ws[f"F{row}"].value).lower().strip()
                    region = nz_str(src_ws[f"H{row}"].value).lower().strip()
                    asym_raw = src_ws[f"S{row}"].value
                    
                    # Skip if no movement/region or no asymmetry data
                    if not movement or not region:
                        continue
                    if test_type == 'lower' and region == "trunk":  # Trunk handled separately with force calculation
                        continue
                    if asym_raw in (None, ""):
                        continue
                    
                    # Try to parse asymmetry - only include if valid
                    pct_value, side_char = parse_asymmetry(asym_raw)
                    if pct_value is None:
                        continue
                    
                    key = (movement, region)
                    if key not in movements_present:
                        movements_present[key] = []
                    movements_present[key].append(row)
            
            # Special handling for trunk - calculate asymmetry from force values
            if test_type == 'lower':
                trunk_pct, trunk_weak_side = calculate_trunk_asymmetry(rows, src_ws)
                if trunk_pct is not None:
                    print(f"    Calculated trunk asymmetry: {trunk_pct:.1f}% (weak side: {trunk_weak_side})")
                    # Store trunk data
                    patient_data['cells']['AA21'] = "Trunk Lateral Flexion /\nالثني الجانبي للجذع"
                    patient_data['cells']['AB21'] = trunk_pct / 100  # Convert to decimal for Excel
                    if trunk_weak_side == "Right":
                        patient_data['cells']['AA22'] = "Right Sides /\nالجانب الأيمن"
                    else:
                        patient_data['cells']['AA22'] = "Left Sides /\nالجانب الأيسر"
                    
                    # Track trunk movement for body parts detection
                    patient_data['movements'].append(('lateral flexion', 'trunk'))
                
                # Assign knee movements to C21/C23 dynamically
                knee_movements = []
                if ('extension', 'knee') in movements_present:
                    knee_movements.append(('extension', 'knee'))
                if ('flexion', 'knee') in movements_present:
                    knee_movements.append(('flexion', 'knee'))
                
                # Assign hip abd/add movements to O21/O23 dynamically
                hip_abd_add_movements = []
                if ('adduction', 'hip') in movements_present:
                    hip_abd_add_movements.append(('adduction', 'hip'))
                if ('abduction', 'hip') in movements_present:
                    hip_abd_add_movements.append(('abduction', 'hip'))
                
                # Assign hip flex/ext movements to AG21/AG23 dynamically
                hip_flex_ext_movements = []
                if ('flexion', 'hip') in movements_present:
                    hip_flex_ext_movements.append(('flexion', 'hip'))
                if ('extension', 'hip') in movements_present:
                    hip_flex_ext_movements.append(('extension', 'hip'))
            
            # Prepare movement lists for upper body
            if test_type == 'upper':
                # Shoulder rotations/movements in C21/C23/C25/C27
                shoulder_c_movements = []
                if ('external rotation', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('external rotation', 'shoulder'))
                if ('internal rotation', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('internal rotation', 'shoulder'))
                if ('flexion', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('flexion', 'shoulder'))
                if ('abduction', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('abduction', 'shoulder'))
                
                # Shoulder push/pull in O21/O23
                shoulder_o_movements = []
                if ('push', 'shoulder') in movements_present:
                    shoulder_o_movements.append(('push', 'shoulder'))
                if ('pull', 'shoulder') in movements_present:
                    shoulder_o_movements.append(('pull', 'shoulder'))
                
                # Elbow in AA21/AA23
                elbow_movements = []
                if ('extension', 'elbow') in movements_present:
                    elbow_movements.append(('extension', 'elbow'))
                if ('flexion', 'elbow') in movements_present:
                    elbow_movements.append(('flexion', 'elbow'))
                
                # Hand/Grip in AG21 only
                hand_movements = []
                if ('grip squeeze', 'hand') in movements_present:
                    hand_movements.append(('grip squeeze', 'hand'))
            
            # Prepare movement lists for full body
            if test_type == 'full':
                # Shoulder rotations/movements in C21/C23/C25/C27
                shoulder_c_movements = []
                if ('external rotation', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('external rotation', 'shoulder'))
                if ('internal rotation', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('internal rotation', 'shoulder'))
                if ('flexion', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('flexion', 'shoulder'))
                if ('abduction', 'shoulder') in movements_present:
                    shoulder_c_movements.append(('abduction', 'shoulder'))
                
                # Elbow in O21/O23
                elbow_o_movements = []
                if ('extension', 'elbow') in movements_present:
                    elbow_o_movements.append(('extension', 'elbow'))
                if ('flexion', 'elbow') in movements_present:
                    elbow_o_movements.append(('flexion', 'elbow'))
                
                # Knee in AA21/AA23
                knee_aa_movements = []
                if ('extension', 'knee') in movements_present:
                    knee_aa_movements.append(('extension', 'knee'))
                if ('flexion', 'knee') in movements_present:
                    knee_aa_movements.append(('flexion', 'knee'))
                
                # Hip abd/add in AG21/AG23
                hip_ag_movements = []
                if ('abduction', 'hip') in movements_present:
                    hip_ag_movements.append(('abduction', 'hip'))
                if ('adduction', 'hip') in movements_present:
                    hip_ag_movements.append(('adduction', 'hip'))

            # For each row belonging to this patient
            for row in rows:
                date_val = src_ws[f"C{row}"].value
                movement = nz_str(src_ws[f"F{row}"].value)
                region = nz_str(src_ws[f"H{row}"].value)
                asym_raw = src_ws[f"S{row}"].value

                # Date - first non-empty for this patient
                if not date_set and date_val not in (None, ""):
                    patient_data['date'] = date_val
                    date_set = True

                if not movement.strip() or not region.strip():
                    continue
                
                # Skip trunk for normal processing since we handle it specially
                if region.lower().strip() == "trunk":
                    continue
                
                if asym_raw in (None, ""):
                    continue
                
                # Filter rows by test type if multiple test types exist
                if len(test_types) > 1:
                    row_test_type = get_movement_test_type(movement, region)
                    # Skip if this movement doesn't belong to current test type
                    if row_test_type and row_test_type != test_type:
                        print(f"    Skipping {region} {movement} (belongs to {row_test_type}, processing {test_type})")
                        continue

                pct_value, side_char = parse_asymmetry(asym_raw)
                if pct_value is None:
                    continue

                # Special handling for 0% asymmetry
                if pct_value == 0:
                    pct_value = 0.1
                    weak_side = None  # Leave weak side empty
                else:
                    if side_char is None:
                        continue
                    
                    side_char = side_char.upper()
                    if side_char == "L":
                        weak_side = "Right"
                    elif side_char == "R":
                        weak_side = "Left"
                    else:
                        continue

                # Use the appropriate mapping function based on body type
                if test_type == 'lower':
                    # Lower body uses dynamic cell assignment
                    label_cell, pct_cell_addr, side_cell_addr, right_text, left_text = get_lower_body_cells(
                        movement, region, knee_movements, hip_abd_add_movements, hip_flex_ext_movements
                    )
                elif test_type == 'upper':
                    # Upper body uses dynamic cell assignment
                    label_cell, pct_cell_addr, side_cell_addr, right_text, left_text = get_upper_body_cells(
                        movement, region, shoulder_c_movements, shoulder_o_movements, elbow_movements, hand_movements
                    )
                elif test_type == 'full':
                    # Full body uses dynamic cell assignment
                    label_cell, pct_cell_addr, side_cell_addr, right_text, left_text = get_full_body_cells(
                        movement, region, shoulder_c_movements, elbow_o_movements, knee_aa_movements, hip_ag_movements
                    )
                else:
                    # Fallback to None if unknown test type
                    label_cell = None
                    pct_cell_addr, side_cell_addr, right_text, left_text = None, None, None, None
                
                if pct_cell_addr is None:
                    continue

                # Get movement label
                movement_label = get_movement_label(movement, region)

                # Keep only largest absolute asymmetry
                existing = patient_data['cells'].get(pct_cell_addr)
                if existing is None:
                    should_update = True
                else:
                    should_update = abs(pct_value) > abs(nz_float(existing))

                if should_update:
                    # Store movement label if we have a label cell (lower body)
                    if label_cell and movement_label:
                        patient_data['cells'][label_cell] = movement_label
                    
                    # Store numeric value divided by 100 (Excel will format as percentage)
                    patient_data['cells'][pct_cell_addr] = pct_value / 100
                    
                    # Only set weak side if it's not None (i.e., not a 0% asymmetry case)
                    if weak_side is not None:
                        if weak_side == "Right":
                            patient_data['cells'][side_cell_addr] = right_text
                        else:
                            patient_data['cells'][side_cell_addr] = left_text
                    
                    # Track this movement for body parts detection
                    patient_data['movements'].append((movement.lower().strip(), region.lower().strip()))

            # Normalize patient name - remove extra spaces
            normalized_patient_name = re.sub(r'\s+', ' ', patient_name).strip()
            
            # Create subfolder based on gym (Body Masters or Body Motions)
            gym_subfolder = os.path.join(programs_folder, gym_folder)
            os.makedirs(gym_subfolder, exist_ok=True)
            
            # Determine output path for this patient
            # Always include test type in filename
            safe_name = make_safe_filename(f"{normalized_patient_name} - {test_type.capitalize()} Body")
            out_path = os.path.join(gym_subfolder, safe_name)
            
            # Use xlwings to fill template
            success = fill_template_with_xlwings(template_path, out_path, patient_name, patient_data, gym_folder)
            
            if success:
                print(f"    Saved: {out_path}")
                
                # Log this test with movement count
                test_date = patient_data.get('date')
                movement_count = len(patient_data.get('movements', []))
                if test_date:
                    log_test(gym_folder, base_dir, patient_name, test_type, test_date, movement_count)
            else:
                print(f"    Failed to save: {out_path}")

    print("All reports generated.")


if __name__ == "__main__":
    main()
